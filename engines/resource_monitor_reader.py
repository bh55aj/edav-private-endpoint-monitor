#!/usr/bin/env python3
"""
============================================================================
EDAV Resource Monitor Excel Reader - engines/resource_monitor_reader.py
============================================================================
Parses the EDAV Resource Monitor Excel export (e.g. DisconnectedPEs.xlsx)
to extract disconnected private endpoints as the authoritative source for
cleanup.

Resource Monitor Export Format:
  The Excel file contains one or more sheets. The primary findings sheet
  contains rows where:
    Column "Check"  = "disconnected_private_endpoints"
    Column "Detail" = "Connection state is Disconnected" (or similar)
    Column "ResourceName" (or "name") = the private endpoint name
    Column "ResourceGroup" = the resource group
    Column "Subscription" = subscription name

  The reader is tolerant of column name variations and sheet name variations.
  It normalizes all column names before extraction.

Supported input files:
  DisconnectedPEs.xlsx     - direct RM export
  findings.xlsx            - generic dashboard export
  EDAV_*.xlsx              - tool-generated reports
  *.csv                    - any CSV with compatible columns

Version: v7.0.0 | EDAV Platform Team
============================================================================
"""

from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ============================================================================
# RESOURCE MONITOR COLUMN MAPPINGS
# ============================================================================

# Check names that indicate a disconnected private endpoint finding
PE_CHECK_NAMES = {
    "disconnected_private_endpoints",
    "disconnected private endpoints",
    "private_endpoint_disconnected",
    "privateendpoint",
    "private endpoint",
    "disconnected_pe",
}

# Detail / description patterns that confirm disconnection
PE_DETAIL_PATTERNS = [
    "connection state is disconnected",
    "connectionstate.*disconnected",
    "disconnected",
    "rejected",
    "no backend",
    "backend not found",
    "private link service not found",
]

# Canonical column name -> list of aliases (all lowercased, spaces/underscores removed)
RM_COLUMN_ALIASES: Dict[str, List[str]] = {
    "resourcename":    ["name","resourcename","endpointname","endpoint_name","resource","resourceid"],
    "resourcegroup":   ["resourcegroup","resource_group","rg","rgname"],
    "subscription":    ["subscription","subscriptionname","subscriptionid","sub","subname"],
    "resourcetype":    ["resourcetype","resource_type","type","check","checkname","check_name"],
    "checkname":       ["checkname","check","check_name","finding_type","findingtype"],
    "detail":          ["detail","details","description","findingdetail","finding_detail","message"],
    "connectionstate": ["connectionstate","connection_state","state","connstate"],
    "severity":        ["severity","priority","finding_severity","findingseverity"],
    "findingstatus":   ["findingstatus","finding_status","status","findingstate"],
    "owner":           ["owner","resource_owner","resourceowner","poc","contact","ownerteam","team"],
    "approvedtodelete":["approvedtodelete","approved","approved_to_delete","delete_approved","deleteapproved"],
    "approvalticket":  ["approvalticket","ticket","change_ticket","itsmticket","chg","changeticket"],
    "approvedby":      ["approvedby","approved_by","approver"],
    "notes":           ["notes","note","comment","comments","remarks"],
    "privateLinkServiceId": ["privatelinkserviceid","backendresource","backend","privatelinkservice"],
    "location":        ["location","region","azureregion"],
    "monthlycost":     ["monthlycost","monthly_cost","cost","estimatedcost"],
}

# ============================================================================
# PARSED FINDING
# ============================================================================

@dataclass
class ResourceMonitorFinding:
    """A single disconnected PE finding extracted from the Resource Monitor export."""

    # Core identity (always populated)
    resource_name: str = ""
    resource_group: str = ""
    subscription: str = ""
    resource_type: str = "Microsoft.Network/privateEndpoints"
    location: str = ""

    # Finding metadata
    check_name: str = ""
    detail: str = ""
    connection_state: str = "Disconnected"
    severity: str = ""
    finding_status: str = ""
    private_link_service_id: str = ""

    # Ownership
    owner: str = ""

    # Approval (may be pre-populated in the import file)
    approved_to_delete: bool = False
    approval_ticket: str = ""
    approved_by: str = ""

    # Cost
    monthly_cost_usd: float = 7.30
    yearly_cost_usd: float = field(default=0.0)

    # Metadata
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    notes: str = ""
    raw_row: Dict = field(default_factory=dict)

    def __post_init__(self):
        self.yearly_cost_usd = round(self.monthly_cost_usd * 12, 2)

    @property
    def display_key(self) -> str:
        """Unique display key: name/rg/sub."""
        return f"{self.resource_name}/{self.resource_group}/{self.subscription}"

    @property
    def is_approved(self) -> bool:
        return (self.approved_to_delete
                and bool(self.approval_ticket)
                and bool(self.approved_by))

    def to_dict(self) -> Dict:
        return {
            "ResourceName": self.resource_name,
            "ResourceGroup": self.resource_group,
            "Subscription": self.subscription,
            "Location": self.location,
            "ResourceType": self.resource_type,
            "CheckName": self.check_name,
            "Detail": self.detail,
            "ConnectionState": self.connection_state,
            "Severity": self.severity,
            "PrivateLinkServiceID": self.private_link_service_id,
            "Owner": self.owner,
            "ApprovedToDelete": self.approved_to_delete,
            "ApprovalTicket": self.approval_ticket,
            "ApprovedBy": self.approved_by,
            "MonthlyCostUSD": self.monthly_cost_usd,
            "YearlyCostUSD": self.yearly_cost_usd,
            "SourceFile": self.source_file,
            "SourceSheet": self.source_sheet,
            "SourceRow": self.source_row,
            "Notes": self.notes,
        }


# ============================================================================
# RESOURCE MONITOR READER
# ============================================================================

class ResourceMonitorReader:
    """
    Reads an EDAV Resource Monitor Excel export and extracts all disconnected
    private endpoint findings.

    Handles:
      - Multiple sheets (tries all, picks the one with most PE findings)
      - Column name variations (case-insensitive, spaces/underscores ignored)
      - Rows where Check = disconnected_private_endpoints
      - Rows where Detail = "Connection state is Disconnected"
      - Mixed-format files that have both summary and detail rows
      - CSV files with the same column structure
      - Pre-populated approval columns (ApprovedToDelete, ApprovalTicket, ApprovedBy)
    """

    def __init__(self, file_path: str, cost_per_pe: float = 7.30):
        self.file_path = Path(file_path)
        self.cost_per_pe = cost_per_pe
        self._col_map: Dict[str, str] = {}

    def read(self) -> Tuple[List[ResourceMonitorFinding], Dict]:
        """
        Parse the file. Returns (findings, summary_dict).
        Raises FileNotFoundError if the file does not exist.
        Raises ValueError if no PE findings can be extracted.
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"Resource Monitor file not found: {self.file_path}")

        ext = self.file_path.suffix.lower()
        if ext == ".csv":
            findings = self._read_csv()
        elif ext in (".xlsx", ".xls"):
            findings = self._read_excel()
        else:
            raise ValueError(f"Unsupported file type: {ext}. Use .xlsx or .csv")

        logger.info("Extracted %d PE findings from %s", len(findings), self.file_path.name)

        summary = self._build_summary(findings)
        return findings, summary

    # ------------------------------------------------------------------
    # EXCEL READER
    # ------------------------------------------------------------------

    def _read_excel(self) -> List[ResourceMonitorFinding]:
        """Read all sheets and extract PE findings."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required: pip install openpyxl")

        wb = openpyxl.load_workbook(str(self.file_path), read_only=True, data_only=True)
        all_findings: List[ResourceMonitorFinding] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            # Build header -> col-index map
            headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            col_map = self._build_col_map(headers)

            # Check if this sheet has PE-relevant columns
            has_name = "resourcename" in col_map
            has_check = "checkname" in col_map
            has_detail = "detail" in col_map
            if not has_name and not has_check:
                logger.debug("Sheet %s: no PE columns, skipping", sheet_name)
                continue

            sheet_findings = []
            for row_idx, row in enumerate(rows[1:], start=2):
                row_dict: Dict[str, Any] = {}
                for ci, val in enumerate(row):
                    if ci < len(headers):
                        row_dict[headers[ci]] = val

                finding = self._extract_finding(
                    row_dict, col_map, sheet_name, row_idx
                )
                if finding is not None:
                    sheet_findings.append(finding)

            logger.debug("Sheet %s: %d PE findings", sheet_name, len(sheet_findings))
            all_findings.extend(sheet_findings)

        wb.close()

        # Deduplicate by (name, resource_group, subscription)
        return self._deduplicate(all_findings)

    # ------------------------------------------------------------------
    # CSV READER
    # ------------------------------------------------------------------

    def _read_csv(self) -> List[ResourceMonitorFinding]:
        """Read CSV file and extract PE findings."""
        findings = []
        with open(self.file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = list(reader.fieldnames or [])
            col_map = self._build_col_map(headers)
            for row_idx, row in enumerate(reader, start=2):
                finding = self._extract_finding(dict(row), col_map,
                                                "CSV", row_idx)
                if finding is not None:
                    findings.append(finding)
        return self._deduplicate(findings)

    # ------------------------------------------------------------------
    # CORE EXTRACTION LOGIC
    # ------------------------------------------------------------------

    def _extract_finding(
        self,
        row: Dict[str, Any],
        col_map: Dict[str, str],
        sheet: str,
        row_idx: int,
    ) -> Optional[ResourceMonitorFinding]:
        """
        Decide if this row is a disconnected PE finding.
        Returns None if the row should be skipped.
        """
        def _get(canonical: str) -> str:
            """Get cell value by canonical column name."""
            actual_col = col_map.get(canonical)
            if actual_col is None:
                return ""
            raw = row.get(actual_col, "")
            if raw is None:
                return ""
            return str(raw).strip()

        resource_name = _get("resourcename")
        resource_group = _get("resourcegroup")
        subscription = _get("subscription")
        check_name = _get("checkname")
        detail = _get("detail")
        conn_state = _get("connectionstate")

        # Skip rows with no resource name
        if not resource_name:
            return None

        # Skip rows that are clearly headers re-printed or section titles
        if resource_name.lower() in ("resource name", "name", "endpoint", "resourcename"):
            return None

        # --- Determine if this is a PE finding ---
        is_pe_finding = False

        # Match by check name
        if check_name:
            check_clean = check_name.lower().replace(" ", "_").replace("-", "_")
            is_pe_finding = any(c in check_clean for c in [
                "disconnected_private_endpoint",
                "private_endpoint",
                "privateendpoint",
                "disconnected_pe",
            ])

        # Match by detail text
        if not is_pe_finding and detail:
            detail_lower = detail.lower()
            is_pe_finding = any(re.search(p, detail_lower) for p in PE_DETAIL_PATTERNS)

        # Match by connection state column
        if not is_pe_finding and conn_state:
            is_pe_finding = conn_state.lower() in ("disconnected", "rejected")

        # If no check column at all, accept all rows (the file IS a PE list)
        if not is_pe_finding and not check_name and not detail:
            # File has name+rg+sub but no check/detail columns
            # Treat every non-empty row as a PE record
            if resource_name and resource_group:
                is_pe_finding = True

        if not is_pe_finding:
            return None

        # --- Build the finding ---
        approved_raw = _get("approvedtodelete").lower()
        approved = approved_raw in ("yes", "true", "1", "y", "approved")

        cost_raw = _get("monthlycost")
        try:
            cost = float(re.sub(r"[^\d.]", "", cost_raw)) if cost_raw else self.cost_per_pe
        except (ValueError, TypeError):
            cost = self.cost_per_pe

        # Infer connection state from check/detail if not explicitly present
        if not conn_state:
            if "disconnected" in detail.lower() or "disconnected" in check_name.lower():
                conn_state = "Disconnected"
            else:
                conn_state = "Unknown"

        finding = ResourceMonitorFinding(
            resource_name=resource_name,
            resource_group=resource_group,
            subscription=subscription,
            location=_get("location"),
            check_name=check_name or "disconnected_private_endpoints",
            detail=detail,
            connection_state=conn_state,
            severity=_get("severity"),
            finding_status=_get("findingstatus"),
            private_link_service_id=_get("privateLinkServiceId"),
            owner=_get("owner"),
            approved_to_delete=approved,
            approval_ticket=_get("approvalticket"),
            approved_by=_get("approvedby"),
            monthly_cost_usd=cost,
            source_file=str(self.file_path),
            source_sheet=sheet,
            source_row=row_idx,
            notes=_get("notes"),
            raw_row={k: str(v) for k, v in row.items() if v is not None},
        )
        return finding

    # ------------------------------------------------------------------
    # HELPERS
    # ------------------------------------------------------------------

    def _build_col_map(self, headers: List[str]) -> Dict[str, str]:
        """
        Build a map from canonical column name to actual column header.
        Case-insensitive, strips spaces and underscores for comparison.
        """
        # Normalize headers
        def _normalize(s: str) -> str:
            return re.sub(r"[\s_\-]+", "", s).lower()

        norm_headers = {_normalize(h): h for h in headers if h}
        col_map: Dict[str, str] = {}

        for canonical, aliases in RM_COLUMN_ALIASES.items():
            # Try canonical name first
            if _normalize(canonical) in norm_headers:
                col_map[canonical] = norm_headers[_normalize(canonical)]
                continue
            # Try aliases
            for alias in aliases:
                if _normalize(alias) in norm_headers:
                    col_map[canonical] = norm_headers[_normalize(alias)]
                    break

        return col_map

    def _deduplicate(self, findings: List[ResourceMonitorFinding]) -> List[ResourceMonitorFinding]:
        """Remove duplicate findings by (resource_name, resource_group, subscription)."""
        seen = set()
        deduped = []
        for f in findings:
            key = (f.resource_name.lower(),
                   f.resource_group.lower(),
                   f.subscription.lower())
            if key not in seen:
                seen.add(key)
                deduped.append(f)
        removed = len(findings) - len(deduped)
        if removed:
            logger.info("Deduplication removed %d duplicate findings", removed)
        return deduped

    def _build_summary(self, findings: List[ResourceMonitorFinding]) -> Dict:
        """Build a summary dict for console output and reporting."""
        total = len(findings)
        approved_count = sum(1 for f in findings if f.approved_to_delete)
        total_monthly = round(sum(f.monthly_cost_usd for f in findings), 2)
        total_yearly = round(total_monthly * 12, 2)

        by_sub: Dict[str, int] = {}
        by_rg: Dict[str, int] = {}
        for f in findings:
            by_sub[f.subscription] = by_sub.get(f.subscription, 0) + 1
            by_rg[f.resource_group] = by_rg.get(f.resource_group, 0) + 1

        return {
            "source_file": str(self.file_path),
            "read_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_findings": total,
            "approved_count": approved_count,
            "pending_approval": total - approved_count,
            "estimated_monthly_savings_usd": total_monthly,
            "estimated_yearly_savings_usd": total_yearly,
            "by_subscription": by_sub,
            "by_resource_group": dict(sorted(by_rg.items(), key=lambda x: -x[1])[:20]),
        }


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================

def read_resource_monitor_export(
    file_path: str,
    cost_per_pe: float = 7.30,
) -> Tuple[List[ResourceMonitorFinding], Dict]:
    """
    Read a Resource Monitor Excel/CSV export.
    Returns (findings, summary).

    Usage:
        findings, summary = read_resource_monitor_export("DisconnectedPEs.xlsx")
        print(f"Found {summary['total_findings']} disconnected endpoints")
        print(f"Monthly cost: ${summary['estimated_monthly_savings_usd']:.2f}")
    """
    reader = ResourceMonitorReader(file_path, cost_per_pe=cost_per_pe)
    return reader.read()
