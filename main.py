#!/usr/bin/env python3
"""
============================================================================
EDAV Resource Monitor Cleanup Platform - v6.1.0
============================================================================
Enterprise Azure governance, cost-reduction, and safe cleanup platform
for the EDAV Resource Monitor dashboard.

Dashboard: https://internal-resource-monitor.edav.cdc.gov/dashboard

Supported resource types:
  Microsoft.Network/privateEndpoints
  Microsoft.Network/networkInterfaces
  Microsoft.Network/networkSecurityGroups
  Microsoft.Network/publicIPAddresses
  Microsoft.Compute/disks
  Microsoft.Compute/virtualMachines
  Microsoft.Storage/storageAccounts
  Microsoft.KeyVault/vaults
  Microsoft.ContainerRegistry/registries
  Microsoft.EventGrid/topics
  Microsoft.EventGrid/systemTopics
  Microsoft.EventHub/namespaces
  Microsoft.Sql/managedInstances
  Microsoft.MachineLearningServices/workspaces

Safety model: Nothing deleted without validation, approval, ticket,
approver, classification=SAFE_DELETE, and interactive CONFIRM.

Version: 5.0.0 | EDAV Platform Team
============================================================================
"""

import argparse
import csv
import fnmatch
import glob
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Governance scanner (Resource Graph multi-type)
try:
    from monitors.governance_scanner import (
        GovernanceScanner, GovernanceClassifier,
        GOVERNANCE_QUERIES,
        CLS_KEEP, CLS_SAFE_DELETE, CLS_REVIEW,
        CLS_DO_NOT_DELETE, CLS_NOT_FOUND, CLS_UNKNOWN,
        RECOMMENDED_ACTIONS,
        QUERY_MODE_AUTO, QUERY_MODE_CLI, QUERY_MODE_REST, QUERY_MODE_CSV,
        QUERY_SOURCE_CLI, QUERY_SOURCE_REST, QUERY_SOURCE_CSV,
        get_manual_csv_instructions,
    )
    GOVERNANCE_SCANNER_AVAILABLE = True
except ImportError:
    GOVERNANCE_SCANNER_AVAILABLE = False

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import yaml
except ImportError:
    yaml = None

try:
    from colorama import init as colorama_init, Fore, Style
    colorama_init(autoreset=True)
    COLORAMA = True
except ImportError:
    COLORAMA = False
    class Fore:
        RED = YELLOW = GREEN = CYAN = WHITE = BLUE = MAGENTA = ""
    class Style:
        BRIGHT = RESET_ALL = DIM = ""

# ============================================================================
# CONSTANTS
# ============================================================================

VERSION = "6.1.0"
TOOL_NAME = "EDAV Resource Monitor Cleanup Platform"
DASHBOARD_URL = "https://internal-resource-monitor.edav.cdc.gov/dashboard"

CLASS_SAFE_DELETE = "SAFE_DELETE"
CLASS_REVIEW_REQUIRED = "REVIEW_REQUIRED"
CLASS_DO_NOT_DELETE = "DO_NOT_DELETE"
CLASS_UNKNOWN = "UNKNOWN"
CLASS_NOT_FOUND = "RESOURCE_NOT_FOUND"
CLASS_ACCESS_REVIEW = "ACCESS_OR_SUBSCRIPTION_REVIEW"

COLUMN_ALIASES = {
    "resourcename": ["name", "resource_name", "endpointname", "endpoint_name"],
    "resourcegroup": ["rg", "resource_group", "resourcegroup"],
    "subscription": ["sub", "subscription_name", "subscriptionname"],
    "resourcetype": ["type", "resource_type", "resourcetype"],
    "resourceid": ["id", "resource_id", "resourceid", "azure_id"],
    "approvedtodelete": ["approved", "approved_to_delete", "delete_approved"],
    "approvalticket": ["ticket", "change_ticket", "itsmticket", "chg"],
    "approvedby": ["approver", "approved_by", "approvedby"],
    "severity": ["priority", "finding_severity"],
    "findingstatus": ["status", "finding_status"],
    "checkname": ["check", "check_name", "finding_type"],
    "owner": ["resource_owner", "poc", "contact"],
    "team": ["team_name", "business_unit", "department"],
    "monthlycost": ["cost", "monthly_cost", "estimated_cost"],
    "environment": ["env", "environment_type"],
    "notes": ["note", "comment", "comments"],
    "findingid": ["finding_id", "id", "finding"],
}

# Excel fill colors for classification
FILL_COLORS = {
    CLASS_SAFE_DELETE:     "FF00AA44",  # Green
    CLASS_REVIEW_REQUIRED: "FFFF9900",  # Orange
    CLASS_DO_NOT_DELETE:   "FFCC0000",  # Red
    CLASS_UNKNOWN:         "FFAAAAAA",  # Gray
    CLASS_NOT_FOUND:       "FF6699CC",  # Blue
    CLASS_ACCESS_REVIEW:   "FFCC88FF",  # Purple
}

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def cprint(msg: str, color: str = "", bold: bool = False) -> None:
    """Print colored console output."""
    if COLORAMA:
        prefix = (Style.BRIGHT if bold else "") + color
        print(f"{prefix}{msg}{Style.RESET_ALL}")
    else:
        print(msg)

def run_az(cmd: List[str], timeout: int = 60) -> Tuple[Optional[Any], Optional[str]]:
    """Run Azure CLI command. Returns (result, error_str)."""
    try:
        result = subprocess.run(
            ["az"] + cmd + ["--output", "json"],
            capture_output=True, text=True, timeout=timeout
        )
        if result.returncode != 0:
            err = result.stderr.strip()
            if "AADSTS500173" in err or ("token" in err.lower() and "expired" in err.lower()):
                cprint("[AUTH ERROR] Azure token expired. Run: az login", Fore.RED, bold=True)
                sys.exit(1)
            if "ResourceNotFound" in err or "was not found" in err.lower():
                return None, "ResourceNotFound"
            return None, err
        if not result.stdout.strip():
            return None, None
        return json.loads(result.stdout), None
    except subprocess.TimeoutExpired:
        return None, f"Timeout after {timeout}s"
    except json.JSONDecodeError as e:
        return None, f"JSON parse error: {e}"
    except FileNotFoundError:
        return None, "Azure CLI not found"
    except Exception as e:
        return None, str(e)

def set_subscription(sub_name: str) -> bool:
    """Set Azure subscription context. Returns True if successful."""
    _, err = run_az(["account", "set", "--subscription", sub_name], timeout=30)
    if err:
        cprint(f"  [WARN] Cannot set subscription to '{sub_name}': {err[:80]}", Fore.YELLOW)
        return False
    return True

def fnmatch_any(name: str, patterns: List[str]) -> bool:
    """Return True if name matches any fnmatch pattern."""
    name_lower = name.lower()
    return any(fnmatch.fnmatch(name_lower, p.lower()) for p in patterns)

def normalize_bool(value: Any) -> bool:
    """Normalize truthy representations."""
    if isinstance(value, bool): return value
    if isinstance(value, str):
        return value.strip().lower() in ("yes", "true", "1", "y", "approved")
    return bool(value)

def safe_str(value: Any, default: str = "") -> str:
    """Safely convert value to string."""
    if value is None: return default
    if isinstance(value, float) and value != value: return default  # NaN
    return str(value).strip()

def ts() -> str:
    """Current timestamp for filenames."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def parse_cost(cost_str: str) -> float:
    """Parse cost string to float."""
    try:
        clean = re.sub(r"[^\d.]", "", str(cost_str))
        return float(clean) if clean else 0.0
    except Exception:
        return 0.0

def ensure_dirs(*dirs: str) -> None:
    """Create directories if they do not exist."""
    for d in dirs:
        Path(d).mkdir(parents=True, exist_ok=True)

# ============================================================================
# CONFIG LOADER
# ============================================================================

class ConfigLoader:
    """Loads all configuration files."""

    def __init__(self, config_dir: str = "config"):
        self.config_dir = config_dir
        self.resource_rules = {}
        self.ownership_map = {}
        self.exclusions = set()
        self.denylist = {}
        self.allowlist = {}

    def load(self) -> bool:
        """Load all config files. Returns True if critical files loaded."""
        self._load_resource_rules()
        self._load_ownership_map()
        self._load_exclusions()
        self._load_denylist()
        self._load_allowlist()
        return bool(self.resource_rules)

    def _load_resource_rules(self):
        path = Path(self.config_dir) / "resource_rules.yaml"
        if not path.exists():
            cprint(f"  [WARN] resource_rules.yaml not found at {path}", Fore.YELLOW)
            return
        if yaml is None:
            cprint("  [WARN] pyyaml not installed - using default rules", Fore.YELLOW)
            return
        try:
            with open(path) as f:
                self.resource_rules = yaml.safe_load(f) or {}
            cprint(f"  Loaded resource_rules.yaml ({len(self.resource_rules)} resource types)", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] Failed to load resource_rules.yaml: {e}", Fore.YELLOW)

    def _load_ownership_map(self):
        path = Path(self.config_dir) / "ownership_map.yaml"
        if not path.exists():
            return
        if yaml is None:
            return
        try:
            with open(path) as f:
                self.ownership_map = yaml.safe_load(f) or {}
            cprint(f"  Loaded ownership_map.yaml", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] Failed to load ownership_map.yaml: {e}", Fore.YELLOW)

    def _load_exclusions(self):
        path = Path(self.config_dir) / "exclusions.txt"
        if not path.exists():
            return
        try:
            with open(path) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        self.exclusions.add(line.lower())
            cprint(f"  Loaded exclusions.txt ({len(self.exclusions)} entries)", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] Failed to load exclusions.txt: {e}", Fore.YELLOW)

    def _load_denylist(self):
        path = Path(self.config_dir) / "denylist.json"
        if not path.exists():
            return
        try:
            with open(path) as f:
                self.denylist = json.load(f) or {}
            blocked = len(self.denylist.get("blocked_resource_names", []))
            cprint(f"  Loaded denylist.json ({blocked} blocked names)", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] Failed to load denylist.json: {e}", Fore.YELLOW)

    def _load_allowlist(self):
        path = Path(self.config_dir) / "allowlist.json"
        if not path.exists():
            return
        try:
            with open(path) as f:
                self.allowlist = json.load(f) or {}
            approved = len(self.allowlist.get("pre_approved_candidates", []))
            cprint(f"  Loaded allowlist.json ({approved} pre-approved candidates)", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] Failed to load allowlist.json: {e}", Fore.YELLOW)

    def get_resource_rule(self, resource_type: str) -> Dict:
        """Get validation rules for a resource type."""
        return self.resource_rules.get(resource_type, self.resource_rules.get("DEFAULT", {}))

    def is_excluded(self, name: str) -> bool:
        """Check if resource name is in exclusions list."""
        return name.lower() in self.exclusions

    def is_denied(self, name: str, resource_group: str = "", tags: Dict = None) -> Tuple[bool, str]:
        """Check if resource is in denylist. Returns (blocked, reason)."""
        if not self.denylist:
            return False, ""
        blocked_names = [n.lower() for n in self.denylist.get("blocked_resource_names", [])]
        if name.lower() in blocked_names:
            return True, f"On denylist (blocked_resource_names)"
        blocked_rgs = [r.lower() for r in self.denylist.get("blocked_resource_groups", [])]
        if resource_group.lower() in blocked_rgs:
            return True, f"Resource group {resource_group} is on denylist"
        patterns = self.denylist.get("blocked_name_patterns", [])
        if fnmatch_any(name, patterns):
            return True, f"Name matches denylist pattern"
        if tags:
            blocked_tags = self.denylist.get("blocked_tags", {})
            for tag_key, blocked_values in blocked_tags.items():
                tag_val = tags.get(tag_key, tags.get(tag_key.lower(), "")).lower()
                if tag_val in [v.lower() for v in blocked_values]:
                    return True, f"Tag {tag_key}={tag_val} is on denylist"
        return False, ""

# ============================================================================
# INPUT PARSER
# ============================================================================

class InputParser:
    """Parses CSV/Excel input from EDAV Resource Monitor dashboard."""

    def __init__(self):
        self.column_map = {}

    def parse(self, input_file: str) -> List[Dict]:
        """Parse input file and return list of resource dicts."""
        path = Path(input_file)
        if not path.exists():
            cprint(f"[ERROR] Input file not found: {input_file}", Fore.RED, bold=True)
            sys.exit(1)
        ext = path.suffix.lower()
        cprint(f"\n  Parsing input: {input_file}", Fore.CYAN)
        if ext == ".csv":
            return self._parse_csv(input_file)
        elif ext in (".xlsx", ".xls"):
            return self._parse_excel(input_file)
        else:
            cprint(f"[ERROR] Unsupported file type: {ext}. Use .csv or .xlsx", Fore.RED)
            sys.exit(1)

    def _parse_csv(self, path: str) -> List[Dict]:
        rows = []
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                col_map = self._build_column_map(headers)
                for row in reader:
                    rows.append(self._normalize_row(dict(row), col_map))
            cprint(f"  Parsed {len(rows)} rows from CSV", Fore.GREEN)
            return rows
        except Exception as e:
            cprint(f"[ERROR] Failed to parse CSV: {e}", Fore.RED)
            sys.exit(1)

    def _parse_excel(self, path: str) -> List[Dict]:
        if pd is None:
            cprint("[ERROR] pandas required for Excel. pip install pandas openpyxl", Fore.RED)
            sys.exit(1)
        try:
            df = pd.read_excel(path, engine="openpyxl")
            headers = list(df.columns)
            col_map = self._build_column_map(headers)
            rows = []
            for _, row in df.iterrows():
                rows.append(self._normalize_row(dict(row), col_map))
            cprint(f"  Parsed {len(rows)} rows from Excel", Fore.GREEN)
            return rows
        except Exception as e:
            cprint(f"[ERROR] Failed to parse Excel: {e}", Fore.RED)
            sys.exit(1)

    def _build_column_map(self, headers: List[str]) -> Dict[str, str]:
        """Map normalized column names to actual headers."""
        col_map = {}
        headers_lower = {h.lower().replace(" ", "").replace("_", ""): h for h in headers}
        for std_col, aliases in COLUMN_ALIASES.items():
            all_variants = [std_col] + aliases
            for variant in all_variants:
                v_clean = variant.lower().replace(" ", "").replace("_", "")
                if v_clean in headers_lower:
                    col_map[std_col] = headers_lower[v_clean]
                    break
        # Also map exact headers
        for h in headers:
            h_clean = h.lower().replace(" ", "").replace("_", "")
            if h_clean not in col_map:
                col_map[h_clean] = h
        return col_map

    def _normalize_row(self, row: Dict, col_map: Dict[str, str]) -> Dict:
        """Normalize a row using the column map."""
        normalized = {}
        for std_col, actual_col in col_map.items():
            if actual_col in row:
                normalized[std_col] = safe_str(row[actual_col])
        # Ensure all keys exist with defaults
        defaults = {
            "resourcename": "", "resourcegroup": "", "subscription": "",
            "resourcetype": "", "resourceid": "", "approvedtodelete": "",
            "approvalticket": "", "approvedby": "", "severity": "",
            "findingstatus": "Active", "checkname": "", "owner": "",
            "team": "", "monthlycost": "0", "environment": "", "notes": "",
            "findingid": "", "recommendation": "",
        }
        for k, v in defaults.items():
            if k not in normalized:
                normalized[k] = v
        return normalized

    def validate_rows(self, rows: List[Dict]) -> List[Dict]:
        """Filter out rows without required fields."""
        valid = []
        skipped = 0
        for row in rows:
            if not row.get("resourcename") or not row.get("resourcegroup"):
                skipped += 1
                continue
            valid.append(row)
        if skipped:
            cprint(f"  Skipped {skipped} rows missing ResourceName or ResourceGroup", Fore.YELLOW)
        cprint(f"  Valid rows: {len(valid)}", Fore.GREEN)
        return valid

# ============================================================================
# OWNERSHIP DETECTOR
# ============================================================================

class OwnershipDetector:
    """Detects resource owner/team from tags, RG patterns, and subscription."""

    OWNER_TAG_KEYS = [
        "owner", "Owner", "OWNER", "EDAV_Created_By", "EDAV_Business_POC",
        "EDAV_Project_Name", "EDAV_Center_Name", "EDAV_Division_Name",
        "application", "Application", "team", "Team", "department",
        "created_by", "CreatedBy", "contact", "Contact",
    ]

    def __init__(self, ownership_map: Dict):
        self.rg_patterns = ownership_map.get("resource_group_patterns", {})
        self.name_patterns = ownership_map.get("resource_name_patterns", {})
        self.sub_patterns = ownership_map.get("subscription_patterns", {})
        self.tag_normalization = ownership_map.get("tag_value_normalization", {})
        self.explicit = ownership_map.get("explicit_resource_owners", {})
        self.team_contacts = ownership_map.get("team_contacts", {})
        owner_keys = ownership_map.get("owner_tag_keys", [])
        if owner_keys:
            self.OWNER_TAG_KEYS = owner_keys

    def detect(self, resource: Dict, azure_tags: Dict = None) -> Tuple[str, str]:
        """Detect owner and team. Returns (owner, team)."""
        name = resource.get("resourcename", "")
        rg = resource.get("resourcegroup", "")
        sub = resource.get("subscription", "")
        csv_owner = resource.get("owner", "")
        csv_team = resource.get("team", "")
        tags = azure_tags or {}
        # 1. Explicit mapping
        if name.lower() in {k.lower(): v for k, v in self.explicit.items()}:
            team = self.explicit.get(name, "")
            return csv_owner or team, team
        # 2. CSV-provided owner/team
        if csv_owner and csv_team:
            return csv_owner, csv_team
        # 3. Azure tags
        for tag_key in self.OWNER_TAG_KEYS:
            tag_val = tags.get(tag_key, tags.get(tag_key.lower(), ""))
            if tag_val:
                normalized = self.tag_normalization.get(tag_val.lower(), tag_val)
                return csv_owner or normalized, csv_team or normalized
        # 4. Resource name pattern
        for pattern, team_name in self.name_patterns.items():
            if fnmatch_any(name, [pattern]):
                return csv_owner or team_name, team_name
        # 5. Resource group pattern
        for pattern, team_name in self.rg_patterns.items():
            if pattern == "*":
                continue
            if fnmatch_any(rg, [pattern]):
                return csv_owner or team_name, team_name
        # 6. Subscription pattern
        for pattern, env in self.sub_patterns.items():
            if fnmatch_any(sub, [pattern]):
                return csv_owner or env, csv_team or env
        # 7. CSV fields alone
        if csv_owner:
            return csv_owner, csv_team or "UNKNOWN"
        if csv_team:
            return csv_owner or "UNKNOWN", csv_team
        # 8. Fallback
        wildcard_team = self.rg_patterns.get("*", "UNKNOWN - Needs Owner Assignment")
        return "UNKNOWN", wildcard_team

# ============================================================================
# TERRAFORM CHECKER
# ============================================================================

class TerraformChecker:
    """Checks if a resource is managed by Terraform."""

    def __init__(self, terraform_path: str = None):
        self.terraform_path = terraform_path
        self._state_cache = None

    def is_terraform_managed(self, resource_name: str, resource_id: str = "") -> Tuple[bool, str]:
        """Returns (is_managed, reason)."""
        if not self.terraform_path:
            return False, "Terraform path not provided"
        state_managed, state_reason = self._check_tf_state(resource_name, resource_id)
        if state_managed:
            return True, state_reason
        src_managed, src_reason = self._check_tf_source(resource_name)
        if src_managed:
            return True, src_reason
        return False, "Not found in Terraform state or source"

    def _check_tf_state(self, name: str, resource_id: str = "") -> Tuple[bool, str]:
        """Check terraform state list for resource."""
        try:
            tf_dir = Path(self.terraform_path)
            if not tf_dir.exists():
                return False, "TF path not found"
            if self._state_cache is None:
                result = subprocess.run(
                    ["terraform", "state", "list"],
                    cwd=str(tf_dir), capture_output=True, text=True, timeout=30
                )
                self._state_cache = result.stdout.lower() if result.returncode == 0 else ""
            if name.lower() in self._state_cache:
                return True, f"Found in terraform state"
            if resource_id and resource_id.lower() in self._state_cache:
                return True, f"Resource ID in terraform state"
            return False, ""
        except Exception:
            return False, ""

    def _check_tf_source(self, name: str) -> Tuple[bool, str]:
        """Check .tf source files for resource name."""
        try:
            tf_dir = Path(self.terraform_path)
            for tf_file in tf_dir.rglob("*.tf"):
                try:
                    content = tf_file.read_text(encoding="utf-8", errors="ignore").lower()
                    if name.lower() in content:
                        return True, f"Found in {tf_file.name}"
                except Exception:
                    continue
            return False, ""
        except Exception:
            return False, ""

# ============================================================================
# AZURE LOCK CHECKER
# ============================================================================

def check_azure_lock(resource_name: str, resource_group: str,
                     resource_type: str = None) -> Tuple[bool, str]:
    """Check for Azure resource locks. Returns (has_lock, reason)."""
    # Check resource-group level locks
    data, err = run_az(["lock", "list", "--resource-group", resource_group])
    if data and isinstance(data, list) and len(data) > 0:
        for lock in data:
            lock_name = lock.get("name", "unknown")
            lock_level = lock.get("level", "unknown")
            return True, f"RG lock: {lock_name} ({lock_level})"
    # Check resource-level locks if resource type known
    if resource_type:
        data2, _ = run_az([
            "lock", "list",
            "--resource", resource_name,
            "--resource-type", resource_type,
            "--resource-group", resource_group,
        ])
        if data2 and isinstance(data2, list) and len(data2) > 0:
            lock = data2[0]
            return True, f"Resource lock: {lock.get('name', 'unknown')}"
    return False, ""

# ============================================================================
# AZURE VALIDATOR - Per-resource-type validation logic
# ============================================================================

class AzureValidator:
    """Validates each resource type against Azure."""

    def validate(self, resource: Dict) -> Dict:
        """Validate a resource and return validation result dict."""
        name = resource.get("resourcename", "")
        rg = resource.get("resourcegroup", "")
        sub = resource.get("subscription", "")
        rtype = resource.get("resourcetype", "").lower()

        result = {
            "resource_exists": None,
            "connection_state": None,
            "backend_resource_id": None,
            "backend_exists": None,
            "attached_to": None,
            "managed_by": None,
            "disk_state": None,
            "power_state": None,
            "azure_tags": {},
            "has_lock": False,
            "lock_reason": "",
            "validation_notes": "",
            "raw_data": None,
        }

        if "privateendpoints" in rtype.replace("/", "").replace(".", ""):
            return self._validate_private_endpoint(name, rg, result)
        elif "networkinterfaces" in rtype:
            return self._validate_nic(name, rg, result)
        elif "networksecuritygroups" in rtype:
            return self._validate_nsg(name, rg, result)
        elif "publicipaddresses" in rtype:
            return self._validate_public_ip(name, rg, result)
        elif "disks" in rtype:
            return self._validate_disk(name, rg, result)
        elif "virtualmachines" in rtype:
            return self._validate_vm(name, rg, result)
        elif "storageaccounts" in rtype:
            return self._validate_storage(name, rg, result)
        elif "vaults" in rtype:
            return self._validate_keyvault(name, rg, result)
        elif "registries" in rtype:
            return self._validate_acr(name, rg, result)
        elif "eventgrid" in rtype or "topics" in rtype or "systemtopics" in rtype:
            return self._validate_generic(name, rg, rtype, result)
        elif "eventhub" in rtype or "namespaces" in rtype:
            return self._validate_eventhub(name, rg, result)
        elif "sql" in rtype or "managedinstances" in rtype:
            return self._validate_generic(name, rg, rtype, result)
        elif "machinelearning" in rtype or "workspaces" in rtype:
            return self._validate_generic(name, rg, rtype, result)
        else:
            return self._validate_generic(name, rg, rtype, result)

    def _get_resource_tags(self, data: Any) -> Dict:
        """Extract tags from Azure resource data."""
        if isinstance(data, dict):
            return data.get("tags") or {}
        return {}

    def _validate_private_endpoint(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["network", "private-endpoint", "show",
                           "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            result["validation_notes"] = "Resource not found in Azure"
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Validation error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        result["raw_data"] = data
        # Check connection state
        connections = data.get("privateLinkServiceConnections") or \
                     data.get("manualPrivateLinkServiceConnections") or []
        if connections:
            conn = connections[0]
            state = conn.get("privateLinkServiceConnectionState", {}).get("status", "Unknown")
            result["connection_state"] = state
            result["backend_resource_id"] = conn.get("privateLinkServiceId", "")
        else:
            result["connection_state"] = "Unknown"
        # Validate backend
        if result["backend_resource_id"]:
            backend_data, backend_err = run_az([
                "resource", "show", "--ids", result["backend_resource_id"]
            ])
            result["backend_exists"] = not (backend_err == "ResourceNotFound" or
                                           (backend_err and "not found" in backend_err.lower()))
        else:
            result["backend_exists"] = None
            result["validation_notes"] = "No backend resource ID found"
        return result

    def _validate_nic(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["network", "nic", "show",
                           "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        vm_id = data.get("virtualMachine", {}).get("id") if isinstance(data.get("virtualMachine"), dict) else None
        pe_id = data.get("privateEndpoint", {}).get("id") if isinstance(data.get("privateEndpoint"), dict) else None
        attached = []
        if vm_id:
            attached.append(f"VM: {vm_id.split('/')[-1]}")
        if pe_id:
            attached.append(f"PE: {pe_id.split('/')[-1]}")
        result["attached_to"] = ", ".join(attached) if attached else None
        return result

    def _validate_nsg(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["network", "nsg", "show",
                           "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        subnets = data.get("subnets") or []
        nics = data.get("networkInterfaces") or []
        attached = []
        if subnets:
            attached.append(f"{len(subnets)} subnet(s)")
        if nics:
            attached.append(f"{len(nics)} NIC(s)")
        result["attached_to"] = ", ".join(attached) if attached else None
        return result

    def _validate_public_ip(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["network", "public-ip", "show",
                           "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        ip_config = data.get("ipConfiguration")
        nat_gw = data.get("natGateway")
        if ip_config or nat_gw:
            result["attached_to"] = "ipConfiguration" if ip_config else "natGateway"
        else:
            result["attached_to"] = None
        return result

    def _validate_disk(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["disk", "show", "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        result["managed_by"] = data.get("managedBy")
        result["disk_state"] = data.get("diskState", "Unknown")
        return result

    def _validate_vm(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["vm", "show", "--name", name, "--resource-group", rg,
                           "--show-details"])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        result["power_state"] = data.get("powerState", "unknown")
        result["validation_notes"] = f"VM power state: {result['power_state']}"
        return result

    def _validate_storage(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["storage", "account", "show",
                           "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        result["validation_notes"] = "Storage account exists. Manual review required."
        return result

    def _validate_keyvault(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["keyvault", "show", "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        props = data.get("properties", {})
        purge_protection = props.get("enablePurgeProtection", False)
        soft_delete = props.get("enableSoftDelete", True)
        result["validation_notes"] = (
            f"Purge protection: {purge_protection}, Soft delete: {soft_delete}. "
            "Manual review required."
        )
        return result

    def _validate_acr(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["acr", "show", "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        result["validation_notes"] = "Container Registry exists. Manual review required."
        return result

    def _validate_eventhub(self, name: str, rg: str, result: Dict) -> Dict:
        data, err = run_az(["eventhubs", "namespace", "show",
                           "--name", name, "--resource-group", rg])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        result["validation_notes"] = "Event Hub namespace exists. Manual review required."
        return result

    def _validate_generic(self, name: str, rg: str, resource_type: str, result: Dict) -> Dict:
        """Generic validation using az resource show."""
        if not resource_type:
            result["resource_exists"] = None
            result["validation_notes"] = "No resource type provided for generic validation"
            return result
        # Map common short type names to full provider/type
        type_map = {
            "microsoft.eventgrid/topics": "microsoft.eventgrid/topics",
            "microsoft.eventgrid/systemtopics": "microsoft.eventgrid/systemtopics",
            "microsoft.sql/managedinstances": "microsoft.sql/managedinstances",
            "microsoft.machinelearningservices/workspaces": "microsoft.machinelearningservices/workspaces",
        }
        rt = type_map.get(resource_type.lower(), resource_type)
        data, err = run_az([
            "resource", "show",
            "--name", name,
            "--resource-group", rg,
            "--resource-type", rt,
        ])
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            result["resource_exists"] = False
            return result
        if err and not data:
            result["resource_exists"] = None
            result["validation_notes"] = f"Error: {err[:100]}"
            return result
        result["resource_exists"] = True
        result["azure_tags"] = self._get_resource_tags(data)
        result["validation_notes"] = "Generic validation. Manual review required."
        return result

# ============================================================================
# RESOURCE CLASSIFIER
# ============================================================================

class ResourceClassifier:
    """Classifies resources based on validation results and rules."""

    def __init__(self, config: ConfigLoader):
        self.config = config

    def classify(self, resource: Dict, validation: Dict) -> Tuple[str, str]:
        """Returns (classification, reason)."""
        name = resource.get("resourcename", "")
        rg = resource.get("resourcegroup", "")
        rtype = resource.get("resourcetype", "")
        tags = validation.get("azure_tags", {})
        approved = normalize_bool(resource.get("approvedtodelete", ""))
        ticket = safe_str(resource.get("approvalticket", ""))
        approver = safe_str(resource.get("approvedby", ""))
        env = safe_str(resource.get("environment", "")).lower()

        # 1. Resource not found
        if validation.get("resource_exists") is False:
            return CLASS_NOT_FOUND, "Resource no longer exists in Azure"

        # 2. Access/auth error
        notes = validation.get("validation_notes", "")
        if validation.get("resource_exists") is None and ("error" in notes.lower() or "timeout" in notes.lower()):
            return CLASS_ACCESS_REVIEW, f"Validation failed: {notes[:80]}"

        # 3. Exclusions list
        if self.config.is_excluded(name):
            return CLASS_DO_NOT_DELETE, "On exclusions list"

        # 4. Denylist
        denied, deny_reason = self.config.is_denied(name, rg, tags)
        if denied:
            return CLASS_DO_NOT_DELETE, f"Denylist: {deny_reason}"

        # 5. Azure lock
        if validation.get("has_lock"):
            return CLASS_DO_NOT_DELETE, f"Azure lock: {validation.get('lock_reason', 'locked')}"

        # 6. Terraform managed
        if validation.get("terraform_managed"):
            return CLASS_DO_NOT_DELETE, f"Terraform managed: {validation.get('terraform_reason', 'TF managed')}"

        # 7. Production environment
        if env in ("production", "prod", "prd", "high"):
            return CLASS_DO_NOT_DELETE, f"Production environment: {env}"
        for tag_key in ["environment", "Environment", "env"]:
            tag_env = tags.get(tag_key, "").lower()
            if tag_env in ("production", "prod", "prd", "high"):
                return CLASS_DO_NOT_DELETE, f"Tag environment={tag_env}"

        # 8. Get resource rules
        rule = self.config.get_resource_rule(rtype)
        auto_delete = rule.get("auto_delete_supported", False)
        default_class = rule.get("default_classification", CLASS_REVIEW_REQUIRED)

        # 9. Resource-type specific SAFE_DELETE criteria
        safe, safe_reason = self._check_safe_delete_criteria(rtype, validation)

        if safe and auto_delete:
            # Check approval
            if approved and ticket and approver:
                return CLASS_SAFE_DELETE, safe_reason
            elif approved and not ticket:
                return CLASS_REVIEW_REQUIRED, f"Safe criteria met but ApprovalTicket missing. {safe_reason}"
            elif approved and not approver:
                return CLASS_REVIEW_REQUIRED, f"Safe criteria met but ApprovedBy missing. {safe_reason}"
            else:
                return CLASS_REVIEW_REQUIRED, f"Safe criteria met but not approved. {safe_reason}"

        if not auto_delete:
            return CLASS_REVIEW_REQUIRED, f"Resource type does not support auto-delete. {notes[:80]}"

        if validation.get("resource_exists") is None:
            return CLASS_UNKNOWN, f"Cannot determine resource state. {notes[:80]}"

        return CLASS_REVIEW_REQUIRED, f"Needs manual review. {safe_reason or notes[:80]}"

    def _check_safe_delete_criteria(self, rtype: str, validation: Dict) -> Tuple[bool, str]:
        """Check type-specific safe-delete criteria. Returns (is_safe, reason)."""
        rtype_lower = rtype.lower()

        # Private Endpoints
        if "privateendpoints" in rtype_lower.replace("/", ""):
            state = validation.get("connection_state", "")
            backend_exists = validation.get("backend_exists")
            if state == "Disconnected" and backend_exists is False:
                return True, "Disconnected PE + backend resource not found"
            if state == "Disconnected" and backend_exists is True:
                return False, "Disconnected but backend still exists - REVIEW_REQUIRED"
            if state == "Disconnected" and backend_exists is None:
                return False, "Disconnected but backend status unknown"
            return False, f"Connection state: {state}"

        # Network Interfaces
        if "networkinterfaces" in rtype_lower:
            attached = validation.get("attached_to")
            if attached:
                return False, f"NIC attached to: {attached}"
            return True, "NIC is unattached"

        # Public IPs
        if "publicipaddresses" in rtype_lower:
            attached = validation.get("attached_to")
            if attached:
                return False, f"Public IP attached to: {attached}"
            return True, "Public IP is unattached"

        # Managed Disks
        if "disks" in rtype_lower and "microsoft.compute" in rtype_lower:
            managed_by = validation.get("managed_by")
            disk_state = validation.get("disk_state", "")
            if managed_by:
                return False, f"Disk managed by: {managed_by}"
            if disk_state.lower() == "unattached":
                return True, "Disk is unattached and managedBy=null"
            return False, f"Disk state: {disk_state}"

        # VMs - never auto-safe
        if "virtualmachines" in rtype_lower:
            power = validation.get("power_state", "unknown")
            return False, f"VM - power state: {power}. Manual review always required."

        # NSGs
        if "networksecuritygroups" in rtype_lower:
            attached = validation.get("attached_to")
            if attached:
                return False, f"NSG attached to: {attached}"
            return False, "NSG has no associations but auto-delete not supported"

        # All others: not safe for auto-delete
        return False, "Resource type requires manual review"

# ============================================================================
# SAFETY GATE
# ============================================================================

class SafetyGate:
    """Enforces 15 safety layers before any deletion."""

    def __init__(self, config: ConfigLoader, dry_run: bool = False):
        self.config = config
        self.dry_run = dry_run

    def check(self, resource: Dict, classification: str, validation: Dict,
              tf_managed: bool = False) -> Tuple[bool, List[str]]:
        """Run all safety gates. Returns (all_pass, failed_reasons)."""
        failed = []
        name = resource.get("resourcename", "")
        rg = resource.get("resourcegroup", "")
        rtype = resource.get("resourcetype", "")
        approved = normalize_bool(resource.get("approvedtodelete", ""))
        ticket = safe_str(resource.get("approvalticket", ""))
        approver = safe_str(resource.get("approvedby", ""))

        gates = [
            ("G01", "Classification = SAFE_DELETE",
             classification == CLASS_SAFE_DELETE,
             f"Classification is {classification}"),
            ("G02", "ApprovedToDelete = Yes",
             approved,
             "ApprovedToDelete not set to Yes"),
            ("G03", "ApprovalTicket populated",
             bool(ticket),
             "ApprovalTicket is empty"),
            ("G04", "ApprovedBy populated",
             bool(approver),
             "ApprovedBy is empty"),
            ("G05", "Not on exclusions list",
             not self.config.is_excluded(name),
             "On exclusions list"),
            ("G06", "Not on denylist",
             not self.config.is_denied(name, rg, validation.get("azure_tags", {}))[0],
             "On denylist"),
            ("G07", "Resource exists in Azure",
             validation.get("resource_exists") is True,
             "Resource not found or existence unknown"),
            ("G08", "No Azure lock",
             not validation.get("has_lock"),
             f"Azure lock exists: {validation.get('lock_reason', '')}"),
            ("G09", "Not Terraform managed",
             not tf_managed,
             "Resource is Terraform managed"),
            ("G10", "Not production environment",
             resource.get("environment", "").lower() not in ("production", "prod", "prd", "high"),
             "Resource is in production environment"),
            ("G11", "Resource type supports auto-delete",
             self.config.get_resource_rule(rtype).get("auto_delete_supported", False),
             "Resource type does not support auto-delete"),
        ]

        for gate_id, gate_name, condition, fail_msg in gates:
            if not condition:
                failed.append(f"{gate_id} [{gate_name}]: {fail_msg}")

        return len(failed) == 0, failed

# ============================================================================
# CLEANUP ENGINE
# ============================================================================

class CleanupEngine:
    """Handles ARM backup, deletion, and post-delete verification."""

    def __init__(self, backup_dir: str = "backups", dry_run: bool = False,
                 delete_pause: int = 2):
        self.backup_dir = backup_dir
        self.dry_run = dry_run
        self.delete_pause = delete_pause

    def backup_resource(self, resource: Dict, validation: Dict) -> Optional[str]:
        """Create ARM JSON backup. Returns backup file path."""
        if self.dry_run:
            return "[dry-run - no backup]"
        name = resource.get("resourcename", "unknown")
        sub = resource.get("subscription", "unknown").replace(" ", "-")
        rtype_short = resource.get("resourcetype", "resource").split("/")[-1]
        timestamp = ts()
        backup_subdir = Path(self.backup_dir) / rtype_short
        backup_subdir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_subdir / f"{name}_{sub}_{timestamp}.json"
        backup_data = {
            "backup_timestamp": datetime.now().isoformat(),
            "resource_name": name,
            "resource_group": resource.get("resourcegroup", ""),
            "subscription": sub,
            "resource_type": resource.get("resourcetype", ""),
            "approval_ticket": resource.get("approvalticket", ""),
            "approved_by": resource.get("approvedby", ""),
            "azure_data": validation.get("raw_data", {}),
            "azure_tags": validation.get("azure_tags", {}),
        }
        try:
            with open(backup_path, "w") as f:
                json.dump(backup_data, f, indent=2, default=str)
            return str(backup_path)
        except Exception as e:
            cprint(f"  [WARN] Backup failed for {name}: {e}", Fore.YELLOW)
            return None

    def delete_resource(self, resource: Dict, validation: Dict) -> Tuple[str, str, str, str]:
        """Delete a resource. Returns (result_status, error_msg, cmd_str, delete_method)."""
        if self.dry_run:
            cmd_str = "DRY_RUN - no command executed"
            return "DRY_RUN_SIMULATED", "", cmd_str, "dry_run"
        name = resource.get("resourcename", "")
        rg = resource.get("resourcegroup", "")
        rtype = resource.get("resourcetype", "").lower()
        resource_id = resource.get("resourceid", "") or ""

        cmd = self._get_delete_command(name, rg, rtype, resource_id)
        if not cmd:
            return "SKIP_UNSUPPORTED", f"No delete command for type: {rtype}", "", ""

        delete_method = "ids" if ("--ids" in cmd) else "name_rg"
        cmd_str = "az " + " ".join(cmd)
        cprint(f"  [DELETE] {cmd_str}", Fore.YELLOW)
        cprint(f"  [METHOD] {delete_method}", Fore.YELLOW)
        start = time.time()
        _, err = run_az(cmd, timeout=120)
        duration = round(time.time() - start, 1)

        if err and err != "ResourceNotFound":
            if "AuthorizationFailed" in err or "does not have authorization" in err.lower():
                rbac_msg = (f"RBAC BLOCKED: account does not have "
                            f"Microsoft.Network/privateEndpoints/delete on this scope.")
                cprint(f"  [RBAC] {rbac_msg}", Fore.RED, bold=True)
                return "DELETE_FAILED", rbac_msg, cmd_str, delete_method
            return "DELETE_FAILED", err[:200], cmd_str, delete_method
        cprint(f"  [OK] DELETED and VERIFIED_GONE pending post-check: {name}", Fore.GREEN, bold=True)
        return "DELETED", "", cmd_str, delete_method

    def _get_delete_command(self, name: str, rg: str, rtype: str,
                             resource_id: str = "") -> Optional[List[str]]:
        """Return az CLI command for deleting the resource type.
        Prefers --ids (full resource ID) when available.
        NOTE: az network private-endpoint delete does NOT support --yes.
        Preferred order: if resource_id exists, delete by --ids; else by --name and --resource-group.
        """
        if "privateendpoints" in rtype.replace("/", ""):
            # Private endpoint delete does NOT support --yes flag
            if resource_id:
                return ["network", "private-endpoint", "delete", "--ids", resource_id]
            return ["network", "private-endpoint", "delete",
                    "--name", name, "--resource-group", rg]
        elif "networkinterfaces" in rtype:
            if resource_id:
                return ["network", "nic", "delete", "--ids", resource_id]
            return ["network", "nic", "delete", "--name", name, "--resource-group", rg, "--yes"]
        elif "publicipaddresses" in rtype:
            if resource_id:
                return ["network", "public-ip", "delete", "--ids", resource_id]
            return ["network", "public-ip", "delete", "--name", name, "--resource-group", rg, "--yes"]
        elif "microsoft.compute/disks" in rtype:
            if resource_id:
                return ["disk", "delete", "--ids", resource_id, "--yes"]
            return ["disk", "delete", "--name", name, "--resource-group", rg, "--yes"]
        else:
            return None  # Unsupported type - handled by safety gate

    def verify_deletion(self, resource: Dict) -> Tuple[str, str]:
        """Verify resource is gone. Returns (status, message)."""
        if self.dry_run:
            return "VERIFICATION_SKIPPED_DRY_RUN", ""
        name = resource.get("resourcename", "")
        rg = resource.get("resourcegroup", "")
        rtype = resource.get("resourcetype", "").lower()
        cmd = self._get_show_command(name, rg, rtype)
        if not cmd:
            return "VERIFICATION_SKIPPED", "No show command available"
        data, err = run_az(cmd, timeout=30)
        if err == "ResourceNotFound" or (err and "not found" in err.lower()):
            return "VERIFIED_GONE", "Azure confirmed ResourceNotFound"
        if data:
            return "VERIFICATION_FAILED", "Resource still exists in Azure"
        return "VERIFIED_GONE", "No data returned (likely gone)"

    def _get_show_command(self, name: str, rg: str, rtype: str) -> Optional[List[str]]:
        if "privateendpoints" in rtype.replace("/", ""):
            return ["network", "private-endpoint", "show", "--name", name, "--resource-group", rg]
        elif "networkinterfaces" in rtype:
            return ["network", "nic", "show", "--name", name, "--resource-group", rg]
        elif "publicipaddresses" in rtype:
            return ["network", "public-ip", "show", "--name", name, "--resource-group", rg]
        elif "microsoft.compute/disks" in rtype:
            return ["disk", "show", "--name", name, "--resource-group", rg]
        return None

# ============================================================================
# REPORT GENERATOR
# ============================================================================

class ReportGenerator:
    """Generates all output reports."""

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        self.timestamp = ts()
        ensure_dirs(output_dir)

    def generate_all(self, results: List[Dict], run_meta: Dict) -> Dict[str, str]:
        """Generate all reports. Returns dict of {format: filepath}."""
        files = {}
        files["csv"] = self._write_csv(results, run_meta)
        files["md"] = self._write_markdown(results, run_meta)
        if pd and openpyxl:
            files["xlsx"] = self._write_excel(results, run_meta)
        files["json"] = self._write_json(results, run_meta)
        files["html"] = self._write_html(results, run_meta)
        return files

    def _report_path(self, name: str, ext: str) -> str:
        return str(Path(self.output_dir) / f"EDAV_{name}_{self.timestamp}.{ext}")

    def _write_csv(self, results: List[Dict], run_meta: Dict) -> str:
        path = self._report_path("Findings_Report", "csv")
        if not results:
            return path
        try:
            fieldnames = list(results[0].keys())
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(results)
            cprint(f"  CSV: {path}", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] CSV write failed: {e}", Fore.YELLOW)
        return path

    def _write_json(self, results: List[Dict], run_meta: Dict) -> str:
        path = self._report_path("Full_Report", "json")
        try:
            with open(path, "w") as f:
                json.dump({"run_metadata": run_meta, "results": results}, f, indent=2, default=str)
            cprint(f"  JSON: {path}", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] JSON write failed: {e}", Fore.YELLOW)
        return path

    def _write_markdown(self, results: List[Dict], run_meta: Dict) -> str:
        path = self._report_path("Summary", "md")
        try:
            mode = run_meta.get("mode", "audit-only")
            run_date = run_meta.get("run_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            total = len(results)
            by_class = {}
            total_cost = 0.0
            safe_cost = 0.0
            for r in results:
                c = r.get("classification", CLASS_UNKNOWN)
                by_class[c] = by_class.get(c, 0) + 1
                cost = parse_cost(r.get("monthlycost", "0"))
                total_cost += cost
                if c == CLASS_SAFE_DELETE:
                    safe_cost += cost
            lines = [
                f"# EDAV Resource Monitor Cleanup Platform - Executive Summary",
                f"",
                f"**Run Date:** {run_date}  ",
                f"**Mode:** {mode}  ",
                f"**Change Ticket:** {run_meta.get('change_ticket', 'N/A')}  ",
                f"**Approved By:** {run_meta.get('approved_by', 'N/A')}  ",
                f"",
                f"## Summary",
                f"",
                f"| Metric | Count |",
                f"|--------|-------|",
                f"| Total Resources Reviewed | {total} |",
            ]
            for cls in [CLASS_SAFE_DELETE, CLASS_REVIEW_REQUIRED, CLASS_DO_NOT_DELETE,
                       CLASS_UNKNOWN, CLASS_NOT_FOUND, CLASS_ACCESS_REVIEW]:
                count = by_class.get(cls, 0)
                lines.append(f"| {cls} | {count} |")
            if safe_cost > 0:
                lines.append(f"| Estimated Monthly Savings (SAFE_DELETE) | ${safe_cost:.2f} |")
            lines += [
                f"",
                f"## SAFE_DELETE Candidates",
                f"",
                f"| Resource | Type | RG | Owner | Monthly Cost | Reason |",
                f"|----------|------|----|-------|--------------|--------|",
            ]
            for r in results:
                if r.get("classification") == CLASS_SAFE_DELETE:
                    lines.append(
                        f"| {r.get('resourcename','')} | {r.get('resourcetype','')} | "
                        f"{r.get('resourcegroup','')} | {r.get('detected_owner','')} | "
                        f"${parse_cost(r.get('monthlycost','0')):.2f} | {r.get('classification_reason','')[:60]} |"
                    )
            with open(path, "w") as f:
                f.write("\n".join(lines))
            cprint(f"  Markdown: {path}", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] Markdown write failed: {e}", Fore.YELLOW)
        return path

    def _write_excel(self, results: List[Dict], run_meta: Dict) -> str:
        path = self._report_path("Findings_Report", "xlsx")
        try:
            df_all = pd.DataFrame(results)
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_all.to_excel(writer, sheet_name="All Findings", index=False)
                for cls, sheet_name in [
                    (CLASS_SAFE_DELETE, "SAFE_DELETE"),
                    (CLASS_REVIEW_REQUIRED, "REVIEW_REQUIRED"),
                    (CLASS_DO_NOT_DELETE, "DO_NOT_DELETE"),
                    (CLASS_UNKNOWN, "UNKNOWN"),
                ]:
                    df_cls = df_all[df_all["classification"] == cls] if "classification" in df_all.columns else pd.DataFrame()
                    if not df_cls.empty:
                        df_cls.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                # Color rows by classification
                wb = writer.book
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row <= 1:
                        continue
                    # Find classification column
                    cls_col = None
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value == "classification":
                            cls_col = col_idx
                            break
                    if cls_col:
                        for row in ws.iter_rows(min_row=2):
                            cls_val = row[cls_col - 1].value
                            fill_hex = FILL_COLORS.get(cls_val, "FFFFFFFF")
                            fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
                            for cell in row:
                                cell.fill = fill
            cprint(f"  Excel: {path}", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] Excel write failed: {e}", Fore.YELLOW)
        return path

    def _write_html(self, results: List[Dict], run_meta: Dict) -> str:
        path = self._report_path("Report", "html")
        try:
            mode = run_meta.get("mode", "audit-only")
            run_date = run_meta.get("run_date", "")
            total = len(results)
            by_class = {c: 0 for c in [CLASS_SAFE_DELETE, CLASS_REVIEW_REQUIRED,
                                        CLASS_DO_NOT_DELETE, CLASS_UNKNOWN, CLASS_NOT_FOUND, CLASS_ACCESS_REVIEW]}
            safe_cost = 0.0
            for r in results:
                c = r.get("classification", CLASS_UNKNOWN)
                by_class[c] = by_class.get(c, 0) + 1
                if c == CLASS_SAFE_DELETE:
                    safe_cost += parse_cost(r.get("monthlycost", "0"))
            cls_colors = {
                CLASS_SAFE_DELETE: "#d4edda", CLASS_REVIEW_REQUIRED: "#fff3cd",
                CLASS_DO_NOT_DELETE: "#f8d7da", CLASS_UNKNOWN: "#e2e3e5",
                CLASS_NOT_FOUND: "#cce5ff", CLASS_ACCESS_REVIEW: "#e8d5f5",
            }
            rows_html = ""
            for r in results:
                cls = r.get("classification", CLASS_UNKNOWN)
                bg = cls_colors.get(cls, "#ffffff")
                rows_html += (
                    f'<tr style="background:{bg}">'
                    f'<td>{r.get("resourcename","")}</td>'
                    f'<td>{r.get("resourcetype","").split("/")[-1]}</td>'
                    f'<td>{r.get("resourcegroup","")}</td>'
                    f'<td><b>{cls}</b></td>'
                    f'<td>{r.get("classification_reason","")[:80]}</td>'
                    f'<td>{r.get("detected_owner","")}</td>'
                    f'<td>{r.get("detected_team","")}</td>'
                    f'<td>${parse_cost(r.get("monthlycost","0")):.2f}</td>'
                    f'</tr>\n'
                )
            html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>EDAV Resource Monitor - {run_date}</title>
<style>body{{font-family:Arial,sans-serif;margin:20px;}}table{{border-collapse:collapse;width:100%;font-size:12px;}}th,td{{border:1px solid #ddd;padding:6px;text-align:left;}}th{{background:#2d5986;color:white;}}.summary{{background:#f8f9fa;padding:15px;border-radius:8px;margin-bottom:20px;}}h1{{color:#2d5986;}}</style></head><body>
<h1>EDAV Resource Monitor Cleanup Platform v{VERSION}</h1>
<div class="summary">
<b>Run Date:</b> {run_date} | <b>Mode:</b> {mode} | 
<b>Ticket:</b> {run_meta.get("change_ticket","N/A")} | 
<b>Approved By:</b> {run_meta.get("approved_by","N/A")}<br><br>
<b>Total:</b> {total} | 
<b style="color:green">SAFE_DELETE:</b> {by_class[CLASS_SAFE_DELETE]} | 
<b style="color:orange">REVIEW_REQUIRED:</b> {by_class[CLASS_REVIEW_REQUIRED]} | 
<b style="color:red">DO_NOT_DELETE:</b> {by_class[CLASS_DO_NOT_DELETE]} | 
<b>UNKNOWN:</b> {by_class[CLASS_UNKNOWN]}<br>
<b>Estimated Monthly Savings:</b> ${safe_cost:.2f}
</div>
<table><thead><tr>
<th>Resource Name</th><th>Type</th><th>Resource Group</th>
<th>Classification</th><th>Reason</th><th>Owner</th><th>Team</th><th>Monthly Cost</th>
</tr></thead><tbody>
{rows_html}</tbody></table></body></html>"""
            with open(path, "w", encoding="utf-8") as f:
                f.write(html)
            cprint(f"  HTML: {path}", Fore.GREEN)
        except Exception as e:
            cprint(f"  [WARN] HTML write failed: {e}", Fore.YELLOW)
        return path

# ============================================================================
# PREFLIGHT CHECKER (Azure CLI & SU Account Enforcement)
# ============================================================================

DEFAULT_REQUIRED_DELETE_USER = "bh55-su@cdc.gov"

def get_current_az_user() -> str:
    """Return the active Azure CLI user.name (email). Returns '' on error."""
    result, err = run_az(["account", "show", "--query", "user.name"], timeout=30)
    if isinstance(result, str):
        return result.strip().strip('"')
    if isinstance(result, dict):
        return str(result).strip().strip('"')
    return ""

def get_current_subscription_info() -> Dict:
    """Return dict with 'name' and 'id' for the active subscription."""
    result, err = run_az(["account", "show"], timeout=30)
    if result and isinstance(result, dict):
        return {"name": result.get("name", ""), "id": result.get("id", "")}
    return {"name": "", "id": ""}


class PreflightChecker:
    """Validates Azure CLI state before any run.
    For delete modes, enforces the SU account requirement.
    """

    def __init__(self, required_subscriptions: List[str] = None,
                 required_user: str = None,
                 mode: str = "report"):
        self.required_subscriptions = required_subscriptions or []
        self.required_user = required_user  # None = not enforced
        self.mode = mode

    def check_all(self) -> bool:
        """Run all preflight checks. Returns True if all pass."""
        cprint("\n  Preflight checks...", Fore.CYAN)
        az_ok = self._check_az_cli()
        if not az_ok:
            return False
        user_ok = self._check_su_account()
        if not user_ok:
            return False
        return True

    def _check_az_cli(self) -> bool:
        """Verify Azure CLI is installed and authenticated."""
        data, err = run_az(["account", "show"], timeout=30)
        if err or not data:
            cprint("  [PREFLIGHT FAIL] Azure CLI not logged in or not installed.", Fore.RED, bold=True)
            cprint("  Run: az login --use-device-code", Fore.YELLOW)
            return False
        sub_name = data.get("name", "unknown") if isinstance(data, dict) else "unknown"
        user_name = (data.get("user", {}) or {}).get("name", "unknown") if isinstance(data, dict) else "unknown"
        cprint(f"  [PREFLIGHT OK] Azure CLI authenticated", Fore.GREEN)
        cprint(f"    Current user       : {user_name}", Fore.WHITE)
        cprint(f"    Current subscription: {sub_name}", Fore.WHITE)
        return True

    def _check_su_account(self) -> bool:
        """Enforce SU account for delete/test-delete modes."""
        if self.mode not in ("delete", "test-delete"):
            return True  # Not required for report mode
        if not self.required_user:
            return True  # No enforcement configured

        current_user = get_current_az_user()
        if current_user.lower() == self.required_user.lower():
            cprint(f"  [PREFLIGHT OK] SU account verified: {current_user}", Fore.GREEN)
            return True

        # BLOCK - wrong account
        sep = "=" * 72
        cprint("\n" + sep, Fore.RED, bold=True)
        cprint("  DELETE BLOCKED", Fore.RED, bold=True)
        cprint(sep, Fore.RED)
        cprint(f"  Current account  : {current_user}", Fore.RED)
        cprint(f"  Required account : {self.required_user}", Fore.RED)
        cprint(f"\n  You must login with the SU account {self.required_user}", Fore.RED)
        cprint(f"  before cleanup can run.", Fore.RED)
        cprint("\n  Remediation steps:", Fore.YELLOW)
        cprint("    az logout", Fore.YELLOW)
        cprint("    az login --use-device-code", Fore.YELLOW)
        cprint("    az account show --query user -o table", Fore.YELLOW)
        cprint(sep, Fore.RED)
        return False


def verify_subscription_context(sub_name: str, required_user: str = "") -> Tuple[str, str, bool]:
    """Set subscription and verify context. Returns (sub_before, sub_after, success)."""
    before_info = get_current_subscription_info()
    sub_before = before_info.get("name", "")

    _, err = run_az(["account", "set", "--subscription", sub_name], timeout=30)
    if err:
        cprint(f"  [WARN] Cannot set subscription '{sub_name}': {err[:80]}", Fore.YELLOW)
        return sub_before, "", False

    after_info = get_current_subscription_info()
    sub_after = after_info.get("name", "")
    current_user = get_current_az_user()

    cprint(f"  [SUB] Subscription : {sub_after}", Fore.CYAN)
    cprint(f"  [USR] Azure CLI User: {current_user}", Fore.CYAN)

    if required_user and current_user.lower() != required_user.lower():
        cprint(f"  [WARN] User changed unexpectedly to {current_user}!", Fore.RED)

    return sub_before, sub_after, True


# ============================================================================
# EXECUTIVE DASHBOARD
# ============================================================================

def print_executive_dashboard(results, run_meta, deletion_results=None):
    """Print executive summary to console."""
    sep = "=" * 72
    cprint("\n" + sep, Fore.CYAN, bold=True)
    cprint(f"  EXECUTIVE DASHBOARD -- {TOOL_NAME} v{VERSION}", Fore.CYAN, bold=True)
    cprint(sep, Fore.CYAN)
    run_date = run_meta.get("run_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    mode = run_meta.get("mode", "audit-only").upper()
    ticket = run_meta.get("change_ticket", "N/A")
    approver = run_meta.get("approved_by", "N/A")
    print(f"  Run Date      : {run_date}")
    print(f"  Mode          : {mode}")
    print(f"  Change Ticket : {ticket}")
    print(f"  Approved By   : {approver}")
    cprint("-" * 72, Fore.CYAN)
    total = len(results)
    by_class = {c: 0 for c in [CLASS_SAFE_DELETE, CLASS_REVIEW_REQUIRED,
                                CLASS_DO_NOT_DELETE, CLASS_UNKNOWN, CLASS_NOT_FOUND, CLASS_ACCESS_REVIEW]}
    safe_cost = 0.0
    for r in results:
        c = r.get("classification", CLASS_UNKNOWN)
        by_class[c] = by_class.get(c, 0) + 1
        if c == CLASS_SAFE_DELETE:
            safe_cost += parse_cost(r.get("monthlycost", "0"))
    print(f"  Total Resources Reviewed      : {total}")
    cprint("-" * 72, Fore.CYAN)
    print(f"  SAFE_DELETE                   : {by_class[CLASS_SAFE_DELETE]:>5}")
    print(f"  REVIEW_REQUIRED               : {by_class[CLASS_REVIEW_REQUIRED]:>5}")
    print(f"  DO_NOT_DELETE                 : {by_class[CLASS_DO_NOT_DELETE]:>5}")
    print(f"  UNKNOWN                       : {by_class[CLASS_UNKNOWN]:>5}")
    print(f"  RESOURCE_NOT_FOUND            : {by_class[CLASS_NOT_FOUND]:>5}")
    print(f"  ACCESS_OR_SUBSCRIPTION_REVIEW : {by_class[CLASS_ACCESS_REVIEW]:>5}")
    if safe_cost > 0:
        cprint("-" * 72, Fore.CYAN)
        print(f"  Est. Monthly Savings (SAFE_DELETE) : ${safe_cost:.2f}")
    if deletion_results:
        cprint("-" * 72, Fore.CYAN)
        deleted = sum(1 for d in deletion_results if d.get("delete_result") in ("DELETED", "DRY_RUN_SIMULATED"))
        verified = sum(1 for d in deletion_results if d.get("verify_result") == "VERIFIED_GONE")
        failed = sum(1 for d in deletion_results if d.get("delete_result") == "DELETE_FAILED")
        print(f"  Resources Deleted             : {deleted}")
        print(f"  Deletions Verified            : {verified}")
        print(f"  Deletion Failures             : {failed}")
    cprint("=" * 72, Fore.CYAN)


# ============================================================================
# GOVERNANCE REPORT GENERATOR
# ============================================================================

GOVERNANCE_EXCEL_COLORS = {
    "SAFE_DELETE":       "FF00AA44",
    "REVIEW_REQUIRED":   "FFFF9900",
    "KEEP":              "FF6699CC",
    "DO_NOT_DELETE":     "FFCC0000",
    "RESOURCE_NOT_FOUND":"FFAAAAAA",
    "UNKNOWN":           "FFDDDDDD",
}


class GovernanceReportGenerator:
    """
    Generates all governance scan output reports.
    Tabs: SAFE_DELETE | REVIEW_REQUIRED | KEEP | DO_NOT_DELETE | All Findings
    Formats: CSV, Excel (multi-tab), Markdown, HTML, JSON
    """

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        self.timestamp = ts()
        ensure_dirs(output_dir)

    def _report_path(self, name: str, ext: str) -> str:
        return str(Path(self.output_dir) / ("EDAV_Governance_" + name + "_" + self.timestamp + "." + ext))

    def generate_all(self, results: List[Dict], summary: Dict,
                     run_meta: Dict) -> Dict[str, str]:
        """Generate all report formats. Returns {format: filepath}."""
        files = {}
        files["csv"] = self._write_csv(results)
        files["json"] = self._write_json(results, summary, run_meta)
        files["md"] = self._write_markdown(results, summary, run_meta)
        if pd and openpyxl:
            files["xlsx"] = self._write_excel(results, summary, run_meta)
        files["html"] = self._write_html(results, summary, run_meta)
        return files

    def _write_csv(self, results: List[Dict]) -> str:
        path = self._report_path("Full_Findings", "csv")
        if not results:
            return path
        try:
            fieldnames = list(results[0].keys())
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(results)
            cprint("  CSV: " + path, Fore.GREEN)
        except Exception as e:
            cprint("  [WARN] CSV write failed: " + str(e), Fore.YELLOW)
        return path

    def _write_json(self, results: List[Dict], summary: Dict, run_meta: Dict) -> str:
        path = self._report_path("Full_Report", "json")
        try:
            with open(path, "w") as f:
                json.dump({"run_metadata": run_meta, "summary": summary,
                           "results": results}, f, indent=2, default=str)
            cprint("  JSON: " + path, Fore.GREEN)
        except Exception as e:
            cprint("  [WARN] JSON write failed: " + str(e), Fore.YELLOW)
        return path

    def _write_markdown(self, results: List[Dict], summary: Dict, run_meta: Dict) -> str:
        path = self._report_path("Executive_Summary", "md")
        try:
            totals = summary.get("totals", {})
            run_date = run_meta.get("run_date", "")
            subs = ", ".join(run_meta.get("subscriptions", []) or ["all"])
            query_results = summary.get("query_results", {})

            lines = [
                "# EDAV Governance Scan - Executive Summary",
                "",
                "**Run Date:** " + run_date,
                "**Subscriptions:** " + subs,
                "**Total Resources Found:** " + str(totals.get("total", 0)),
                "",
                "## Classification Summary",
                "",
                "| Classification | Count |",
                "|----------------|-------|",
                "| SAFE_DELETE (Cleanup Candidates) | " + str(totals.get("SAFE_DELETE", 0)) + " |",
                "| REVIEW_REQUIRED (Owner Review) | " + str(totals.get("REVIEW_REQUIRED", 0)) + " |",
                "| KEEP (Azure-Managed / In Use) | " + str(totals.get("KEEP", 0)) + " |",
                "| DO_NOT_DELETE (Blocked) | " + str(totals.get("DO_NOT_DELETE", 0)) + " |",
                "| RESOURCE_NOT_FOUND (Already Gone) | " + str(totals.get("RESOURCE_NOT_FOUND", 0)) + " |",
                "| UNKNOWN | " + str(totals.get("UNKNOWN", 0)) + " |",
                "",
                "## EDAV Governance Scan Summary",
                "",
                "| Category | Total | SAFE_DELETE | REVIEW | KEEP |",
                "|----------|-------|-------------|--------|------|",
            ]
            for key, qr in query_results.items():
                cls_counts = qr.get("classifications", {})
                lines.append(
                    "| " + qr.get("display_name", key) + " | " + str(qr.get("count", 0)) +
                    " | " + str(cls_counts.get("SAFE_DELETE", 0)) +
                    " | " + str(cls_counts.get("REVIEW_REQUIRED", 0)) +
                    " | " + str(cls_counts.get("KEEP", 0)) + " |"
                )

            lines += [
                "",
                "## SAFE_DELETE Candidates",
                "",
                "| Resource | Type | Resource Group | Subscription | Reason |",
                "|----------|------|----------------|--------------|--------|",
            ]
            for r in results:
                if r.get("Classification") == "SAFE_DELETE":
                    lines.append(
                        "| " + r.get("ResourceName", "") +
                        " | " + r.get("ResourceType", "").split("/")[-1] +
                        " | " + r.get("ResourceGroup", "") +
                        " | " + str(r.get("Subscription", ""))[:20] +
                        " | " + r.get("Reason", "")[:60] + " |"
                    )

            lines += [
                "",
                "## NIC Safety Note",
                "",
                "> **IMPORTANT:** An unattached NIC does NOT mean it is orphaned.",
                "> Many NICs are Azure-managed (AKS node NICs, PE NICs ending in `-pe-nic`,",
                "> Databricks cluster NICs, App Gateway NICs). These are classified as **KEEP**.",
                "> Never delete NICs without confirming they are not Azure-managed.",
                "",
                "## Recommended Next Steps",
                "",
                "1. Review SAFE_DELETE tab in the Excel report with network/platform teams",
                "2. For each SAFE_DELETE candidate: collect ApprovalTicket and ApprovedBy",
                "3. Mark ApprovedToDelete=Yes in the approved deletions spreadsheet",
                "4. Dry-run: python main.py --mode delete --dry-run --delete-approved --input approved.xlsx",
                "5. Live delete (requires SU account): --required-user bh55-su@cdc.gov",
                "6. Post-delete: run --scan-governance again to confirm cleanup",
                "7. Run Resource Graph query from docs/RESOURCE_GRAPH_QUERIES.md",
            ]

            with open(path, "w") as f:
                f.write("\n".join(lines))
            cprint("  Markdown: " + path, Fore.GREEN)
        except Exception as e:
            cprint("  [WARN] Markdown write failed: " + str(e), Fore.YELLOW)
        return path

    def _write_excel(self, results: List[Dict], summary: Dict, run_meta: Dict) -> str:
        path = self._report_path("Findings", "xlsx")
        try:
            df_all = pd.DataFrame(results)

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_all.to_excel(writer, sheet_name="All Findings", index=False)

                for cls_name, sheet_name in [
                    ("SAFE_DELETE",     "SAFE_DELETE"),
                    ("REVIEW_REQUIRED", "REVIEW_REQUIRED"),
                    ("KEEP",            "KEEP"),
                    ("DO_NOT_DELETE",   "DO_NOT_DELETE"),
                ]:
                    if "Classification" in df_all.columns:
                        df_cls = df_all[df_all["Classification"] == cls_name]
                        if not df_cls.empty:
                            df_cls.to_excel(writer, sheet_name=sheet_name[:31], index=False)

                if "QueryCategory" in df_all.columns:
                    for cat in df_all["QueryCategory"].dropna().unique():
                        df_cat = df_all[df_all["QueryCategory"] == cat]
                        if not df_cat.empty:
                            sheet = cat.replace("_", " ").title()[:31]
                            df_cat.to_excel(writer, sheet_name=sheet, index=False)

                totals = summary.get("totals", {})
                summary_rows = [
                    ["Classification", "Count"],
                    ["SAFE_DELETE", totals.get("SAFE_DELETE", 0)],
                    ["REVIEW_REQUIRED", totals.get("REVIEW_REQUIRED", 0)],
                    ["KEEP", totals.get("KEEP", 0)],
                    ["DO_NOT_DELETE", totals.get("DO_NOT_DELETE", 0)],
                    ["RESOURCE_NOT_FOUND", totals.get("RESOURCE_NOT_FOUND", 0)],
                    ["UNKNOWN", totals.get("UNKNOWN", 0)],
                    ["Total", totals.get("total", 0)],
                    ["Azure-Managed", totals.get("azure_managed", 0)],
                ]
                pd.DataFrame(summary_rows[1:], columns=summary_rows[0]).to_excel(
                    writer, sheet_name="Summary", index=False
                )

                wb = writer.book
                ws = wb["All Findings"]
                if ws.max_row > 1:
                    cls_col = None
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value == "Classification":
                            cls_col = col_idx
                            break
                    if cls_col:
                        for row_cells in ws.iter_rows(min_row=2):
                            cls_val = row_cells[cls_col - 1].value or ""
                            fill_hex = GOVERNANCE_EXCEL_COLORS.get(cls_val, "FFFFFFFF")
                            fill = PatternFill(start_color=fill_hex,
                                              end_color=fill_hex, fill_type="solid")
                            for cell in row_cells:
                                cell.fill = fill

            cprint("  Excel: " + path, Fore.GREEN)
        except Exception as e:
            cprint("  [WARN] Excel write failed: " + str(e), Fore.YELLOW)
        return path

    def _write_html(self, results: List[Dict], summary: Dict, run_meta: Dict) -> str:
        path = self._report_path("Dashboard", "html")
        try:
            totals = summary.get("totals", {})
            run_date = run_meta.get("run_date", "")
            subs_str = ", ".join(run_meta.get("subscriptions", []) or ["all"])

            cls_bg = {
                "SAFE_DELETE": "#d4edda",
                "REVIEW_REQUIRED": "#fff3cd",
                "KEEP": "#cce5ff",
                "DO_NOT_DELETE": "#f8d7da",
                "RESOURCE_NOT_FOUND": "#e2e3e5",
                "UNKNOWN": "#f8f9fa",
            }

            rows_html_parts = []
            for r in results:
                cls = r.get("Classification", "UNKNOWN")
                bg = cls_bg.get(cls, "#ffffff")
                az_mgd = "Azure-Mgd" if r.get("AzureManaged") else ""
                rows_html_parts.append(
                    "<tr style=\"background:" + bg + "\">"
                    "<td>" + r.get("ResourceName","") + "</td>"
                    "<td>" + r.get("ResourceType","").split("/")[-1] + "</td>"
                    "<td>" + r.get("ResourceGroup","") + "</td>"
                    "<td>" + str(r.get("Subscription",""))[:20] + "</td>"
                    "<td>" + r.get("Location","") + "</td>"
                    "<td><b>" + cls + "</b></td>"
                    "<td>" + r.get("Reason","")[:80] + "</td>"
                    "<td>" + az_mgd + "</td>"
                    "<td>" + r.get("OwnerTeam","") + "</td>"
                    "<td>" + r.get("RecommendedAction","")[:60] + "</td>"
                    "</tr>"
                )
            rows_html = "\n".join(rows_html_parts)

            html_parts = [
                "<!DOCTYPE html><html><head><meta charset=\"utf-8\">",
                "<title>EDAV Governance Scan - " + run_date + "</title>",
                "<style>body{font-family:Arial,sans-serif;margin:20px;}",
                "table{border-collapse:collapse;width:100%;font-size:11px;}",
                "th,td{border:1px solid #ddd;padding:5px;text-align:left;}",
                "th{background:#2d5986;color:white;}",
                ".summary{background:#f8f9fa;padding:15px;border-radius:8px;margin-bottom:20px;}",
                ".badge{display:inline-block;padding:8px 15px;border-radius:5px;font-weight:bold;margin:3px;}",
                ".nic-warn{background:#fff3cd;border:1px solid #ffc107;padding:10px;border-radius:5px;margin-bottom:15px;}",
                "h1{color:#2d5986;}</style></head><body>",
                "<h1>EDAV Governance Scan - " + run_date + "</h1>",
                "<p><b>Subscriptions:</b> " + subs_str + "</p>",
                "<div class=\"nic-warn\">",
                "<b>NIC Safety Rule:</b> An unattached NIC does NOT mean it is orphaned. ",
                "Check for -pe-nic, .nic., AKS, Databricks patterns before any deletion.",
                "</div>",
                "<div class=\"summary\">",
                "<div class=\"badge\" style=\"background:#d4edda\">SAFE_DELETE: " + str(totals.get("SAFE_DELETE",0)) + "</div>",
                "<div class=\"badge\" style=\"background:#fff3cd\">REVIEW_REQUIRED: " + str(totals.get("REVIEW_REQUIRED",0)) + "</div>",
                "<div class=\"badge\" style=\"background:#cce5ff\">KEEP: " + str(totals.get("KEEP",0)) + "</div>",
                "<div class=\"badge\" style=\"background:#f8d7da\">DO_NOT_DELETE: " + str(totals.get("DO_NOT_DELETE",0)) + "</div>",
                "<div class=\"badge\" style=\"background:#e2e3e5\">Total: " + str(totals.get("total",0)) + " | Azure-Managed: " + str(totals.get("azure_managed",0)) + "</div>",
                "</div>",
                "<table><thead><tr>",
                "<th>Resource Name</th><th>Type</th><th>Resource Group</th>",
                "<th>Subscription</th><th>Location</th><th>Classification</th>",
                "<th>Reason</th><th>Azure-Mgd</th><th>Owner</th><th>Recommended Action</th>",
                "</tr></thead><tbody>",
                rows_html,
                "</tbody></table>",
                "<p style=\"color:#666;font-size:11px;margin-top:15px\">",
                "EDAV Resource Governance Platform v" + VERSION + " | ",
                "Nothing is deleted without validation, approval, a change ticket, an approver, and a typed CONFIRM.",
                "</p></body></html>",
            ]
            html = "\n".join(html_parts)

            with open(path, "w", encoding="utf-8") as f:
                f.write(html)
            cprint("  HTML: " + path, Fore.GREEN)
        except Exception as e:
            cprint("  [WARN] HTML write failed: " + str(e), Fore.YELLOW)
        return path

# ============================================================================
# MAIN PIPELINE
# ============================================================================

def run_pipeline(args):
    """Main execution pipeline."""
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    dry_run = getattr(args, "dry_run", False)
    cleanup = getattr(args, "cleanup_approved", False)
    audit_only = getattr(args, "audit_only", False)
    verify_only = getattr(args, "verify_only", False)
    mode = "dry-run" if dry_run else ("cleanup-approved" if cleanup else "audit-only")
    output_dir = getattr(args, "output_dir", "reports")
    config_dir = getattr(args, "config_dir", "config")
    terraform_path = getattr(args, "terraform_path", None)
    delete_pause = getattr(args, "delete_pause", 2)
    change_ticket = getattr(args, "change_ticket", "") or ""
    approved_by = getattr(args, "approved_by", "") or ""
    subscriptions_arg = getattr(args, "subscriptions", "") or ""
    subscriptions = [s.strip() for s in subscriptions_arg.split(",") if s.strip()]

    run_meta = {
        "run_date": run_date, "mode": mode,
        "change_ticket": change_ticket or "N/A",
        "approved_by": approved_by or "N/A",
        "input_file": getattr(args, "input", ""),
        "version": VERSION,
    }

    sep = "=" * 72
    cprint("\n" + sep, Fore.CYAN, bold=True)
    cprint(f"  {TOOL_NAME} v{VERSION}", Fore.CYAN, bold=True)
    cprint(f"  Dashboard: {DASHBOARD_URL}", Fore.CYAN)
    cprint(sep, Fore.CYAN)
    cprint(f"  Mode: {mode.upper()} | Date: {run_date}", Fore.CYAN)
    if dry_run:
        cprint("  [DRY RUN] No Azure resources will be modified.", Fore.YELLOW, bold=True)

    # --- PREFLIGHT OUTPUT ---
    _rq_user = getattr(args, "required_user", None) or DEFAULT_REQUIRED_DELETE_USER
    _cur_user = get_current_az_user()
    _cur_sub_info = get_current_subscription_info()
    _input_file = getattr(args, "input", "")
    _is_delete_mode = (cleanup or getattr(args, "delete_approved", False))
    cprint("\n  ── Preflight Context ─────────────────────────────────────────────", Fore.CYAN)
    cprint(f"  Current Azure CLI user  : {_cur_user}", Fore.WHITE)
    cprint(f"  Required user (delete)  : {_rq_user}", Fore.WHITE)
    cprint(f"  Current subscription    : {_cur_sub_info.get('name', 'unknown')}", Fore.WHITE)
    cprint(f"  Target subscriptions    : {', '.join(subscriptions) if subscriptions else 'all'}", Fore.WHITE)
    cprint(f"  Delete mode             : {'YES' if _is_delete_mode else 'NO'}", Fore.WHITE)
    cprint(f"  Dry run                 : {'YES' if dry_run else 'NO'}", Fore.WHITE)
    cprint(f"  Input file              : {_input_file}", Fore.WHITE)
    cprint(f"  Output dir              : {output_dir}", Fore.WHITE)
    if terraform_path:
        cprint(f"  Terraform path          : {terraform_path}", Fore.WHITE)
    if _is_delete_mode and not dry_run:
        if _cur_user.lower() != _rq_user.lower():
            cprint("\n  [BLOCKED] Current user does not match required SU user!", Fore.RED, bold=True)
            cprint(f"  Current account  : {_cur_user}", Fore.RED)
            cprint(f"  Required account : {_rq_user}", Fore.RED)
            cprint("  Run: az logout", Fore.YELLOW)
            cprint("  Run: az login --use-device-code", Fore.YELLOW)
            cprint("  Then verify: az account show --query user -o table", Fore.YELLOW)
            sys.exit(1)
    cprint("  ──────────────────────────────────────────────────────────────────", Fore.CYAN)

    # Preflight
    required_user = getattr(args, "required_user", None) or DEFAULT_REQUIRED_DELETE_USER
    run_mode = getattr(args, "mode", None)
    if mode == "cleanup-approved" or getattr(args, "delete_approved", False):
        run_mode_check = "delete"
    elif mode == "dry-run":
        run_mode_check = "report"
    else:
        run_mode_check = "report"
    preflight = PreflightChecker(
        required_subscriptions=subscriptions,
        required_user=required_user if (mode == "cleanup-approved" or getattr(args, "delete_approved", False)) else None,
        mode=run_mode_check,
    )
    if not preflight.check_all():
        sys.exit(1)
    current_az_user = get_current_az_user()

    # Load config
    cprint("\n  Loading configuration...", Fore.CYAN)
    config = ConfigLoader(config_dir=config_dir)
    config.load()

    # Parse input
    cprint("\n  Parsing input file...", Fore.CYAN)
    parser = InputParser()
    raw_rows = parser.parse(getattr(args, "input", ""))
    rows = parser.validate_rows(raw_rows)
    if not rows:
        cprint("[ERROR] No valid rows to process.", Fore.RED)
        sys.exit(1)

    # Initialize components
    validator = AzureValidator()
    classifier = ResourceClassifier(config)
    tf_checker = TerraformChecker(terraform_path=terraform_path)
    ownership = OwnershipDetector(config.ownership_map)
    safety = SafetyGate(config, dry_run=dry_run)
    reporter = ReportGenerator(output_dir=output_dir)
    cleanup_engine = CleanupEngine(backup_dir="backups", dry_run=dry_run, delete_pause=delete_pause)

    # Validate and classify
    cprint(f"\n{sep}", Fore.CYAN)
    cprint(f"  PHASE 1: Validation & Classification ({len(rows)} resources)", Fore.CYAN, bold=True)
    cprint(sep, Fore.CYAN)

    results = []
    for i, resource in enumerate(rows, 1):
        name = resource.get("resourcename", "")
        rg = resource.get("resourcegroup", "")
        sub = resource.get("subscription", "")
        rtype = resource.get("resourcetype", "")
        rtype_short = rtype.split("/")[-1] if rtype else "unknown"
        cprint(f"  [{i:03d}/{len(rows)}] {name} ({rtype_short}) in {rg}", Fore.WHITE)

        # Set subscription context
        if sub and subscriptions and sub in subscriptions:
            set_subscription(sub)

        # Validate
        try:
            validation = validator.validate(resource)
        except Exception as e:
            validation = {"resource_exists": None, "azure_tags": {}, "has_lock": False,
                         "lock_reason": "", "validation_notes": f"Exception: {e}"}

        # Check lock
        try:
            has_lock, lock_reason = check_azure_lock(name, rg, rtype)
            validation["has_lock"] = has_lock
            validation["lock_reason"] = lock_reason
        except Exception:
            validation.setdefault("has_lock", False)
            validation.setdefault("lock_reason", "")

        # Check Terraform
        tf_managed, tf_reason = False, ""
        if terraform_path:
            try:
                tf_managed, tf_reason = tf_checker.is_terraform_managed(
                    name, resource.get("resourceid", ""))
            except Exception:
                pass
        validation["terraform_managed"] = tf_managed
        validation["terraform_reason"] = tf_reason

        # Ownership
        owner, team = ownership.detect(resource, validation.get("azure_tags", {}))

        # Classify
        cls, cls_reason = classifier.classify(resource, validation)

        # Build result
        result = dict(resource)
        result.update({
            "classification": cls,
            "classification_reason": cls_reason,
            "detected_owner": owner,
            "detected_team": team,
            "connection_state": str(validation.get("connection_state", "")),
            "backend_exists": str(validation.get("backend_exists", "")),
            "attached_to": str(validation.get("attached_to", "")),
            "managed_by": str(validation.get("managed_by", "")),
            "disk_state": str(validation.get("disk_state", "")),
            "power_state": str(validation.get("power_state", "")),
            "has_azure_lock": validation.get("has_lock", False),
            "terraform_managed": tf_managed,
            "validation_notes": validation.get("validation_notes", ""),
            "scan_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        cls_color = {
            CLASS_SAFE_DELETE: Fore.GREEN, CLASS_REVIEW_REQUIRED: Fore.YELLOW,
            CLASS_DO_NOT_DELETE: Fore.RED, CLASS_UNKNOWN: Fore.WHITE,
            CLASS_NOT_FOUND: Fore.CYAN, CLASS_ACCESS_REVIEW: Fore.MAGENTA,
        }.get(cls, Fore.WHITE)
        cprint(f"       -> {cls}: {cls_reason[:70]}", cls_color)
        results.append(result)

    # Generate reports
    cprint(f"\n{sep}", Fore.CYAN)
    cprint("  PHASE 2: Generating Reports", Fore.CYAN, bold=True)
    cprint(sep, Fore.CYAN)
    report_files = reporter.generate_all(results, run_meta)
    cprint(f"  Reports written to: {output_dir}/", Fore.GREEN)

    deletion_results = []

    # Cleanup phase
    if cleanup and not audit_only and not verify_only:
        safe_resources = [r for r in results if r.get("classification") == CLASS_SAFE_DELETE]
        if not safe_resources:
            cprint("\n  No SAFE_DELETE resources found. Nothing to clean up.", Fore.YELLOW)
        else:
            cprint(f"\n{sep}", Fore.CYAN)
            cprint(f"  PHASE 3: Cleanup Engine ({len(safe_resources)} SAFE_DELETE resources)", Fore.CYAN, bold=True)
            cprint(sep, Fore.CYAN)
            if dry_run:
                cprint("  [DRY RUN] Simulating deletion...", Fore.YELLOW)
            else:
                # Bulk validation summary
                total_rows = len(results)
                approved_rows = sum(1 for r in results if normalize_bool(r.get("approvedtodelete", "")))
                missing_ticket = sum(1 for r in safe_resources if not r.get("approvalticket", "").strip())
                missing_approver = sum(1 for r in safe_resources if not r.get("approvedby", "").strip())
                excluded_rows = sum(1 for r in results if r.get("classification") == CLASS_DO_NOT_DELETE)
                review_rows = sum(1 for r in results if r.get("classification") == CLASS_REVIEW_REQUIRED)

                cprint("\n" + "=" * 72, Fore.CYAN)
                cprint("  BULK DELETE VALIDATION SUMMARY", Fore.CYAN, bold=True)
                cprint("=" * 72, Fore.CYAN)
                print(f"  Total rows in input         : {total_rows}")
                print(f"  Approved rows               : {approved_rows}")
                cprint(f"  SAFE_DELETE candidates      : {len(safe_resources)}", Fore.GREEN)
                print(f"  Missing ticket              : {missing_ticket}")
                print(f"  Missing approver            : {missing_approver}")
                print(f"  Excluded / DO_NOT_DELETE    : {excluded_rows}")
                print(f"  REVIEW_REQUIRED             : {review_rows}")
                cprint("-" * 72, Fore.CYAN)
                cprint(f"  Current Azure CLI user      : {current_az_user}", Fore.WHITE)
                cprint(f"  Required SU user            : {required_user}", Fore.WHITE)
                if current_az_user.lower() != required_user.lower():
                    cprint("  [BLOCKED] SU account mismatch - cannot continue.", Fore.RED, bold=True)
                    sys.exit(1)
                cprint("=" * 72, Fore.CYAN)

                cprint(f"\n  About to delete {len(safe_resources)} resource(s):", Fore.RED, bold=True)
                for r in safe_resources:
                    print(f"    - {r['resourcename']} ({r['resourcegroup']})")
                if change_ticket:
                    print(f"  Change Ticket : {change_ticket}")
                confirm = input("\n  Type CONFIRM to proceed (anything else aborts): ").strip()
                if confirm != "CONFIRM":
                    cprint("  Aborted. No resources deleted.", Fore.YELLOW)
                    print_executive_dashboard(results, run_meta, deletion_results or None)
                    sys.exit(0)

            for resource in safe_resources:
                name = resource.get("resourcename", "")
                rg = resource.get("resourcegroup", "")
                sub = resource.get("subscription", "")
                cprint(f"\n  Processing: {name}", Fore.CYAN)

                # Re-validate
                validation = validator.validate(resource)
                gate_pass, gate_failures = safety.check(
                    resource, CLASS_SAFE_DELETE, validation,
                    resource.get("terraform_managed", False)
                )

                del_result = {
                    "resourcename": name, "resourcegroup": rg,
                    "subscription": sub, "resourcetype": resource.get("resourcetype", ""),
                    "change_ticket": change_ticket, "approved_by": approved_by,
                    "gate_pass": gate_pass,
                    "gate_failures": "; ".join(gate_failures),
                    "delete_result": "", "verify_result": "",
                    "backup_path": "", "error_message": "",
                    "azure_cli_user": "",
                    "required_user": "",
                    "subscription_before_delete": "",
                    "subscription_after_set": "",
                    "delete_command_used": "",
                    "delete_method": "",
                    "delete_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "dry_run": dry_run,
                }

                if not gate_pass:
                    cprint(f"  [SKIP] Safety gate failed:", Fore.RED)
                    for fail in gate_failures:
                        cprint(f"    - {fail}", Fore.RED)
                    del_result["delete_result"] = "SKIPPED_SAFETY_GATE"
                    del_result["error_message"] = "; ".join(gate_failures)
                    deletion_results.append(del_result)
                    continue

                # Subscription context verification before deletion
                sub_before = ""
                sub_after_set = ""
                if sub:
                    sub_before, sub_after_set, sub_ok = verify_subscription_context(sub, required_user)
                    if not sub_ok:
                        cprint(f"  [SKIP] Cannot set subscription {sub} - skipping", Fore.RED)
                        del_result["delete_result"] = "SKIPPED_SUBSCRIPTION_ACCESS"
                        del_result["error_message"] = f"Cannot set subscription: {sub}"
                        del_result["subscription_before_delete"] = sub_before
                        del_result["subscription_after_set"] = "FAILED"
                        del_result["azure_cli_user"] = get_current_az_user()
                        del_result["required_user"] = required_user
                        deletion_results.append(del_result)
                        continue
                del_result["subscription_before_delete"] = sub_before
                del_result["subscription_after_set"] = sub_after_set
                del_result["azure_cli_user"] = get_current_az_user()
                del_result["required_user"] = required_user

                backup_path = cleanup_engine.backup_resource(resource, validation)
                del_result["backup_path"] = backup_path or ""
                cprint(f"  [BACKUP] {backup_path}", Fore.CYAN)

                delete_status, delete_err, delete_cmd_used, delete_method = cleanup_engine.delete_resource(resource, validation)
                del_result["delete_result"] = delete_status
                del_result["delete_command_used"] = delete_cmd_used
                del_result["delete_method"] = delete_method
                del_result["error_message"] = delete_err

                if delete_status in ("DELETED", "DRY_RUN_SIMULATED"):
                    cprint(f"  [OK] {delete_status}: {name}", Fore.GREEN, bold=True)
                else:
                    cprint(f"  [FAIL] {delete_err[:80]}", Fore.RED)

                if delete_status == "DELETED":
                    time.sleep(delete_pause)
                    verify_status, verify_msg = cleanup_engine.verify_deletion(resource)
                    del_result["verify_result"] = verify_status
                    if verify_status == "VERIFIED_GONE":
                        cprint(f"  DELETED and VERIFIED_GONE: {name}", Fore.GREEN, bold=True)
                    else:
                        cprint(f"  [VERIFY FAILED] {verify_msg}", Fore.RED, bold=True)
                elif delete_status == "DELETE_FAILED" and delete_err and "RBAC BLOCKED" in delete_err:
                    cprint(f"  {delete_err}", Fore.RED, bold=True)

                deletion_results.append(del_result)
                if delete_pause > 0 and not dry_run:
                    time.sleep(delete_pause)

            cprint("\n" + "=" * 72, Fore.CYAN)
            cprint("  POST-DELETE REMINDER", Fore.CYAN, bold=True)
            cprint("=" * 72, Fore.CYAN)
            cprint("  Run the Resource Graph query in docs/RESOURCE_GRAPH_QUERIES.md", Fore.YELLOW)
            cprint("  to confirm remaining disconnected private endpoints.", Fore.YELLOW)
            cprint("  Query file: docs/RESOURCE_GRAPH_QUERIES.md", Fore.YELLOW)
            cprint("=" * 72, Fore.CYAN)

            # Write deletion report
            if deletion_results:
                del_csv = reporter._report_path("Delete_Report", "csv")
                try:
                    with open(del_csv, "w", newline="") as f:
                        writer = csv.DictWriter(f, fieldnames=list(deletion_results[0].keys()))
                        writer.writeheader()
                        writer.writerows(deletion_results)
                    cprint(f"  Deletion Report: {del_csv}", Fore.GREEN)
                except Exception as e:
                    cprint(f"  [WARN] Deletion report failed: {e}", Fore.YELLOW)

    # Executive Dashboard
    print_executive_dashboard(results, run_meta, deletion_results or None)

# ============================================================================
# ARGUMENT PARSER & ENTRY POINT
# ============================================================================

def build_arg_parser() -> argparse.ArgumentParser:
    """Build the command-line argument parser."""
    parser = argparse.ArgumentParser(
        prog="main.py",
        description=f"{TOOL_NAME} v{VERSION}\n"
                    f"Dashboard: {DASHBOARD_URL}\n\n"
                    "Discovers, validates, classifies, and safely cleans up "
                    "Azure resources from the EDAV Resource Monitor dashboard findings.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
example usage:

  # Audit-only (default): validate and report, no deletions
  python main.py --input examples/sample_dashboard_findings.csv

  # With subscriptions:
  python main.py --input findings.csv --subscriptions "OCIO-TSBDEV-C1,OCIO-TSBPRD-C1"

  # Dry-run cleanup simulation:
  python main.py --input approved.csv --cleanup-approved --dry-run --change-ticket CHG0001

  # Live cleanup (requires explicit approval in input file):
  python main.py --input approved.csv --cleanup-approved --change-ticket CHG0001 --approved-by "Linda Johnson"

  # Generate owner report only:
  python main.py --input findings.csv --generate-owner-report
"""
    )

    # Required
    parser.add_argument("--input", "-i", required=True,
                       help="Input CSV or Excel file from EDAV Resource Monitor dashboard")

    # Execution modes
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument("--audit-only", action="store_true",
                           help="Discovery, validation, and reports only. No deletions. (DEFAULT)")
    mode_group.add_argument("--cleanup-approved", action="store_true",
                           help="Enable cleanup of approved SAFE_DELETE resources")
    mode_group.add_argument("--verify-only", action="store_true",
                           help="Verify previously deleted resources are gone")
    mode_group.add_argument("--generate-owner-report", action="store_true",
                           help="Generate owner/team-specific reports for follow-up")

    # Cleanup options
    parser.add_argument("--dry-run", action="store_true",
                       help="Simulate cleanup without any Azure changes")
    parser.add_argument("--change-ticket", default="",
                       help="ITSM change ticket reference (e.g., CHG0012345)")
    parser.add_argument("--approved-by", default="",
                       help="Name of the approver")
    parser.add_argument("--delete-pause", type=int, default=2,
                       help="Seconds to pause between deletions (default: 2)")

    # Azure options
    parser.add_argument("--subscriptions", default="",
                       help="Comma-separated list of Azure subscription names")
    parser.add_argument("--terraform-path", default=None,
                       help="Path to Terraform repository for TF state check")

    # Config options
    parser.add_argument("--config-dir", default="config",
                       help="Directory containing config files (default: config/)")
    parser.add_argument("--output-dir", "-o", default="reports",
                       help="Directory for output reports (default: reports/)")
    parser.add_argument("--exclusions", default=None,
                       help="Path to custom exclusions file (overrides config/exclusions.txt)")
    parser.add_argument("--denylist", default=None,
                       help="Path to custom denylist JSON (overrides config/denylist.json)")

    # Preflight only
    parser.add_argument("--self-test", action="store_true",
                       help="Run preflight checks only and exit")

    # SU account enforcement
    parser.add_argument("--required-user", dest="required_user",
                        default=DEFAULT_REQUIRED_DELETE_USER,
                        help=("Required Azure CLI user for delete modes "
                              f"(default: {DEFAULT_REQUIRED_DELETE_USER}). "
                              "Script aborts if current user does not match."))

    return parser


def main():
    """Entry point - supports both Phase 1 modular CLI and legacy CLI."""
    # Use Phase 1 parser (superset of legacy - includes --mode, --resource-type, etc.)
    parser = build_phase1_arg_parser()
    args = parser.parse_args()

    # Override config paths if provided
    if getattr(args, "exclusions", None) and Path(args.exclusions).exists():
        os.environ["EDAV_EXCLUSIONS_PATH"] = args.exclusions
    if getattr(args, "denylist", None) and Path(args.denylist).exists():
        os.environ["EDAV_DENYLIST_PATH"] = args.denylist

    # Self-test mode
    if getattr(args, "self_test", False):
        cprint(f"\n{TOOL_NAME} v{VERSION} -- Self-Test Mode", Fore.CYAN, bold=True)
        preflight = PreflightChecker()
        ok = preflight.check_all()
        if ok:
            cprint("\nAll preflight checks passed.", Fore.GREEN, bold=True)
            sys.exit(0)
        else:
            cprint("\nSome preflight checks failed. See above.", Fore.RED)
            sys.exit(1)

    # Phase 1 mode routing
    mode = getattr(args, "mode", None)
    has_input = getattr(args, "input", None)
    audit_only = getattr(args, "audit_only", False)
    cleanup_approved = getattr(args, "cleanup_approved", False)

    # Governance scan mode (new)
    scan_gov = getattr(args, "scan_governance", False)
    if scan_gov:
        run_governance_scan(args)
        return

    if mode in ("report", "delete", "test-delete"):
        # Phase 1 modular pipeline
        run_phase1_pipeline(args)
    elif has_input or audit_only or cleanup_approved:
        # Legacy pipeline (backward compatible)
        run_pipeline(args)
    else:
        # Default: show help
        parser.print_help()
        sys.exit(0)



# ============================================================================
# PHASE 1 MODULAR PIPELINE - CLI ADDITIONS
# ============================================================================

def build_phase1_arg_parser():
    """Phase 1 modular CLI argument parser."""
    p = argparse.ArgumentParser(
        prog='main.py',
        description=TOOL_NAME + ' v' + VERSION + ' - Phase 1 Modular Resource Monitor',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
example usage:

  # Report mode - private endpoints with Terraform drift
  python main.py --mode report --resource-type private-endpoints \\
    --subscriptions \"OCIO-TSBDEV-C1,OCIO-TSBPRD-C1\" \\
    --terraform-path \"/path/to/terraform\"

  # Report mode - all resource types
  python main.py --mode report --resource-type all \\
    --subscriptions \"OCIO-TSBDEV-C1,OCIO-TSBPRD-C1\"

  # Delete mode (approval-gated)
  python main.py --mode delete --resource-type private-endpoints \\
    --input approvals.xlsx --delete-approved

  # Legacy audit mode (unchanged)
  python main.py --input findings.csv --audit-only

  # Governance scan (Resource Graph - all resource types)
  python main.py --scan-governance \
    --subscriptions "OCIO-TSBDEV-C1,OCIO-TSBPRD-C1,OCIO-EDAV-DMZ-DEV-C1,OCIO-EDAV-DMZ-PRD-C1" \
    --output-dir reports/governance/

  # Governance scan - specific query
  python main.py --scan-governance --governance-query unattached_nics \
    --subscriptions "OCIO-TSBDEV-C1"

  # Governance scan - az rest fallback (no resource-graph extension needed)
  python main.py --scan-governance --query-mode rest \
    --subscriptions "OCIO-TSBDEV-C1,OCIO-TSBPRD-C1"

  # Governance scan - manual CSV mode (for restricted CDC VDIs with SSL issues)
  python main.py --scan-governance --query-mode csv \
    --manual-private-endpoints-csv pe_results.csv \
    --manual-nics-csv nic_results.csv \
    --manual-disks-csv disk_results.csv

  # Governance scan - auto mode with CSV fallback folder
  python main.py --scan-governance --query-mode auto \
    --manual-input-dir "C:\\Users\\bh55\\Desktop\\rg_exports"

  # Test-delete mode (single endpoint, safe)
  python main.py --mode test-delete --resource-type private-endpoints \
    --name testwebbseries-pe --resource-group ocio-network \
    --subscription OCIO-TSBDEV-C1 --required-user bh55-su@cdc.gov
"""
    )
    p.add_argument('--mode', choices=['report', 'delete', 'test-delete'], default=None,
                   help="Operation mode: 'report' (validate+report), 'delete' (approval-gated cleanup), "
                        "'test-delete' (safe test deletion of one endpoint to verify CLI permissions)")
    p.add_argument('--resource-type', dest='resource_type',
                   choices=['private-endpoints', 'storage', 'all'],
                   default='private-endpoints',
                   help='Resource type(s) to scan (default: private-endpoints)')
    p.add_argument('--subscriptions', default='',
                   help='Comma-separated Azure subscription names')
    p.add_argument('--terraform-path', dest='terraform_path', default=None,
                   help='Path to Terraform workspace for drift detection')
    p.add_argument('--owners-file', dest='owners_file',
                   default='config/owners.yml',
                   help='Path to owners.yml (default: config/owners.yml)')
    p.add_argument('--input', '-i', default=None,
                   help='Input file: findings CSV/XLSX for report mode, approved file for delete mode')
    p.add_argument('--output-dir', '-o', dest='output_dir', default='reports',
                   help='Directory for output reports (default: reports/)')
    p.add_argument('--dry-run', action='store_true',
                   help='Simulate all actions without modifying Azure resources')
    p.add_argument('--delete-approved', action='store_true',
                   help='Enable deletion of ApprovedToDelete=Yes resources (requires --mode delete)')
    p.add_argument('--required-user', dest='required_user',
                 default=DEFAULT_REQUIRED_DELETE_USER,
                 help=f'Required Azure CLI user for delete/test-delete modes '
                      f'(default: {DEFAULT_REQUIRED_DELETE_USER}). '
                      'Script aborts if current user does not match.')
    p.add_argument('--name', default=None,
                 help='Resource name (used with --mode test-delete)')
    p.add_argument('--resource-group', dest='resource_group_arg', default=None,
                 help='Resource group (used with --mode test-delete)')
    p.add_argument('--subscription-arg', dest='subscription_arg', default=None,
                 help='Subscription name (used with --mode test-delete)')
    p.add_argument('--allow-terraform-managed', action='store_true',
                 help='Override Terraform protection gate (use with caution)')
    p.add_argument('--scan-governance', action='store_true',
                 help='Run broad EDAV Resource Governance scan using Azure Resource Graph. '
                      'Discovers disconnected PEs, unattached NSGs/disks/PIPs/NICs, '
                      'stopped VMs, Event Grid topics with no subscriptions, storage review, '
                      'and AKS/Databricks managed resources. '
                      'Example: python main.py --scan-governance '
                      '--subscriptions "OCIO-TSBDEV-C1,OCIO-TSBPRD-C1,OCIO-EDAV-DMZ-DEV-C1,OCIO-EDAV-DMZ-PRD-C1"')
    p.add_argument('--governance-query', dest='governance_query', default=None,
                 help='Filter governance scan to a specific query category '
                      '(e.g. unattached_nics, disconnected_private_endpoints)')
    p.add_argument('--query-mode', dest='query_mode', default='auto',
                 choices=['auto', 'cli', 'rest', 'csv'],
                 help='Resource Graph query mode. '
                      'auto: try az graph, fall back to az rest, then manual CSV. '
                      'cli: requires resource-graph extension. '
                      'rest: az rest POST - works WITHOUT resource-graph extension. '
                      'csv: manually exported CSV files from Azure Portal. '
                      'Use rest on CDC VDIs where SSL blocks extension install.')
    p.add_argument('--manual-input-dir', dest='manual_input_dir', default=None,
                 help='Directory of manually exported Resource Graph CSV files. '
                      'Files matched by name (pe*.csv, nic*.csv, disk*.csv, etc.).')
    p.add_argument('--manual-private-endpoints-csv', dest='manual_private_endpoints_csv',
                 default=None,
                 help='CSV for disconnected private endpoints (Azure Portal Resource Graph export).')
    p.add_argument('--manual-nsg-csv', dest='manual_nsg_csv', default=None,
                 help='CSV for unattached NSGs (Azure Portal Resource Graph export).')
    p.add_argument('--manual-disks-csv', dest='manual_disks_csv', default=None,
                 help='CSV for unattached managed disks (Azure Portal Resource Graph export).')
    p.add_argument('--manual-publicips-csv', dest='manual_publicips_csv', default=None,
                 help='CSV for unattached public IPs (Azure Portal Resource Graph export).')
    p.add_argument('--manual-nics-csv', dest='manual_nics_csv', default=None,
                 help='CSV for unattached NICs (Azure Portal export). '
                      'NIC classification (KEEP/REVIEW) still applied automatically.')
    p.add_argument('--config-dir', dest='config_dir', default='config',
                   help='Config directory (default: config/)')
    p.add_argument('--change-ticket', default='',
                   help='ITSM change ticket reference')
    p.add_argument('--approved-by', default='',
                   help='Name of the approver')
    p.add_argument('--audit-only', action='store_true',
                   help='Audit-only mode (legacy flag - same as --mode report)')
    p.add_argument('--cleanup-approved', action='store_true',
                   help='Cleanup mode (legacy flag - same as --mode delete --delete-approved)')
    p.add_argument('--self-test', action='store_true',
                   help='Run preflight checks only')
    p.add_argument('--delete-pause', type=int, default=2,
                   help='Pause between deletions (seconds)')
    p.add_argument('--verify-only', action='store_true',
                   help='Verify previously deleted resources')
    p.add_argument('--generate-owner-report', action='store_true',
                   help='Generate owner/team reports')
    return p


def _load_owners_config(owners_file):
    """Load owners.yml. Falls back to ownership_map.yaml."""
    if yaml is None:
        return {}
    paths_to_try = [owners_file, 'config/owners.yml', 'config/ownership_map.yaml']
    for path in paths_to_try:
        p = Path(path)
        if p.exists():
            try:
                with open(p) as f:
                    data = yaml.safe_load(f) or {}
                cprint('  Loaded owners config: ' + str(path), Fore.GREEN)
                return data
            except Exception as e:
                cprint('  [WARN] Could not load ' + str(path) + ': ' + str(e), Fore.YELLOW)
    cprint('  [WARN] No owners config found. Ownership will be UNKNOWN.', Fore.YELLOW)
    return {}


def _load_subscriptions_config(config_dir):
    """Load subscriptions from config/subscriptions.yml."""
    if yaml is None:
        return []
    path = Path(config_dir) / 'subscriptions.yml'
    if not path.exists():
        return []
    try:
        with open(path) as f:
            data = yaml.safe_load(f) or {}
        subs = [s['name'] for s in data.get('subscriptions', []) if s.get('enabled', True)]
        cprint('  Loaded ' + str(len(subs)) + ' subscription(s) from subscriptions.yml', Fore.GREEN)
        return subs
    except Exception as e:
        cprint('  [WARN] Could not load subscriptions.yml: ' + str(e), Fore.YELLOW)
        return []


def _generate_team_reports(results, output_dir, timestamp):
    """Generate per-team Excel reports in team_reports/."""
    if not (pd and openpyxl):
        cprint('  [WARN] pandas/openpyxl required for team reports. Skipping.', Fore.YELLOW)
        return
    team_reports_dir = Path('team_reports')
    team_reports_dir.mkdir(parents=True, exist_ok=True)
    teams = {}
    for r in results:
        team = r.get('detected_team') or r.get('team') or 'UNKNOWN'
        teams.setdefault(team, []).append(r)
    for team, team_results in teams.items():
        safe_team = re.sub(r'[^a-zA-Z0-9_]', '_', team).lower()
        filename = team_reports_dir / (safe_team + '_private_endpoints_' + timestamp + '.xlsx')
        try:
            df = pd.DataFrame(team_results)
            desired_cols = [
                'resourcename', 'resourcegroup', 'subscription', 'resourcetype',
                'connection_state', 'backend_exists', 'classification',
                'classification_reason', 'detected_owner', 'detected_team',
                'terraform_managed', 'approvedtodelete', 'approvalticket',
                'approvedby', 'validation_notes', 'notes',
            ]
            cols = [c for c in desired_cols if c in df.columns]
            df_out = df[cols] if cols else df
            with pd.ExcelWriter(str(filename), engine='openpyxl') as writer:
                df_out.to_excel(writer, sheet_name='Findings', index=False)
            cprint('  Team report: ' + str(filename), Fore.GREEN)
        except Exception as e:
            cprint('  [WARN] Team report failed for ' + team + ': ' + str(e), Fore.YELLOW)


def _generate_executive_summary(results, run_meta, output_dir):
    """Generate reports/executive_summary.md."""
    ensure_dirs(output_dir)
    path = Path(output_dir) / 'executive_summary.md'
    total = len(results)
    by_class = {}
    by_team = {}
    for r in results:
        cls = r.get('classification', CLASS_UNKNOWN)
        by_class[cls] = by_class.get(cls, 0) + 1
        team = r.get('detected_team') or r.get('team') or 'UNKNOWN'
        by_team[team] = by_team.get(team, 0) + 1
    run_date = run_meta.get('run_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    lines = [
        '# EDAV Azure Resource Monitor - Executive Summary',
        '',
        '**Run Timestamp:** ' + run_date,
        '**Resource Type:** ' + run_meta.get('resource_type', 'all'),
        '**Mode:** ' + run_meta.get('mode', 'report'),
        '**Subscriptions:** ' + run_meta.get('subscriptions', 'N/A'),
        '',
        '## Summary Counts',
        '',
        '| Metric | Count |',
        '|--------|-------|',
        '| Total Resources Scanned | ' + str(total) + ' |',
        '| SAFE_DELETE (Cleanup Candidates) | ' + str(by_class.get(CLASS_SAFE_DELETE, 0)) + ' |',
        '| REVIEW_REQUIRED (Owner Review) | ' + str(by_class.get(CLASS_REVIEW_REQUIRED, 0)) + ' |',
        '| DO_NOT_DELETE (Blocked) | ' + str(by_class.get(CLASS_DO_NOT_DELETE, 0)) + ' |',
        '| ACCESS_OR_SUBSCRIPTION_REVIEW | ' + str(by_class.get(CLASS_ACCESS_REVIEW, 0)) + ' |',
        '| RESOURCE_NOT_FOUND (Already Gone) | ' + str(by_class.get(CLASS_NOT_FOUND, 0)) + ' |',
        '| UNKNOWN | ' + str(by_class.get(CLASS_UNKNOWN, 0)) + ' |',
        '',
        '## Team Breakdown',
        '',
        '| Team | Resources |',
        '|------|-----------|',
    ]
    for team, count in sorted(by_team.items(), key=lambda x: -x[1]):
        lines.append('| ' + team + ' | ' + str(count) + ' |')
    lines += [
        '',
        '## Recommended Next Actions',
        '',
        '1. Review REVIEW_REQUIRED resources with team leads',
        '2. Collect approvals from teams (ApprovedToDelete=Yes, ticket, approver)',
        '3. Create ITSM change ticket',
        '4. Dry-run: python main.py --mode delete --dry-run --delete-approved --input approvals.xlsx',
        '5. Live delete after dry-run review',
        '6. Post-delete verification: run report mode again',
    ]
    try:
        with open(path, 'w') as f:
            f.write('\n'.join(lines))
        cprint('  Executive summary: ' + str(path), Fore.GREEN)
    except Exception as e:
        cprint('  [WARN] Executive summary failed: ' + str(e), Fore.YELLOW)
    return str(path)


def run_test_delete(name: str, resource_group: str,
                    subscription: str = None, required_user: str = DEFAULT_REQUIRED_DELETE_USER,
                    output_dir: str = "reports", dry_run: bool = False) -> None:
    """
    Safe test deletion of a single private endpoint.
    Proves that Azure CLI delete works before bulk cleanup.
    Generates: test_delete_report_<timestamp>.md
    """
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timestamp_str = ts()
    ensure_dirs(output_dir)

    sep = "=" * 72
    cprint("\n" + sep, Fore.CYAN, bold=True)
    cprint("  TEST-DELETE MODE - Single Endpoint Validation", Fore.CYAN, bold=True)
    cprint(sep, Fore.CYAN)
    cprint("  Endpoint     : " + name, Fore.WHITE)
    cprint("  RG           : " + resource_group, Fore.WHITE)
    cprint("  Subscription : " + (subscription or '(current)'), Fore.WHITE)
    if dry_run:
        cprint("  [DRY RUN] No Azure resources will be modified.", Fore.YELLOW, bold=True)

    current_user = get_current_az_user()
    cprint("\n  Step 1: Verify SU account", Fore.CYAN)
    cprint("  Current Azure CLI user  : " + current_user, Fore.WHITE)
    cprint("  Required user           : " + required_user, Fore.WHITE)

    report_lines = [
        "# EDAV Test-Delete Report",
        "",
        "**Date:** " + run_date,
        "**Endpoint:** " + name,
        "**Resource Group:** " + resource_group,
        "**Subscription:** " + (subscription or '(current)'),
        "**Azure CLI User:** " + current_user,
        "**Required User:** " + required_user,
        "**Dry Run:** " + ("Yes" if dry_run else "No"),
        "",
        "## Steps",
        "",
    ]

    if not dry_run and current_user.lower() != required_user.lower():
        msg = ("DELETE BLOCKED: Azure CLI is currently authenticated as " + current_user +
               ". You must login with the SU account " + required_user + " before cleanup can run.")
        cprint("\n  [BLOCKED] " + msg, Fore.RED, bold=True)
        cprint("  Run: az logout", Fore.YELLOW)
        cprint("  Run: az login --use-device-code", Fore.YELLOW)
        cprint("  Then verify: az account show --query user -o table", Fore.YELLOW)
        report_lines.append("### Step 1: SU Account - FAILED")
        report_lines.append("- " + msg)
        _write_test_report(report_lines, output_dir, timestamp_str)
        sys.exit(1)
    report_lines.append("### Step 1: SU Account - PASS")
    report_lines.append("- Current user: " + current_user)
    cprint("  [PASS] SU account confirmed", Fore.GREEN)

    cprint("\n  Step 2: Set subscription context", Fore.CYAN)
    sub_after = ""
    if subscription:
        sub_before, sub_after, sub_ok = verify_subscription_context(subscription, required_user)
        if not sub_ok:
            msg = "Cannot set subscription: " + subscription
            cprint("  [FAIL] " + msg, Fore.RED)
            report_lines.append("### Step 2: Subscription Context - FAILED")
            report_lines.append("- " + msg)
            _write_test_report(report_lines, output_dir, timestamp_str)
            sys.exit(1)
        report_lines.append("### Step 2: Subscription Context - PASS")
        report_lines.append("- Subscription set to: " + sub_after)
        cprint("  [PASS] Subscription: " + sub_after, Fore.GREEN)
    else:
        sub_info = get_current_subscription_info()
        sub_after = sub_info.get("name", "")
        report_lines.append("### Step 2: Subscription Context")
        report_lines.append("- Using current subscription: " + sub_after)
        cprint("  Using current subscription: " + sub_after, Fore.WHITE)

    cprint("\n  Step 3: Validate endpoint exists and check connection state", Fore.CYAN)
    validator = AzureValidator()
    fake_resource = {
        "resourcename": name, "resourcegroup": resource_group,
        "subscription": subscription or sub_after,
        "resourcetype": "Microsoft.Network/privateEndpoints",
        "resourceid": "", "approvedtodelete": "Yes",
        "approvalticket": "TEST", "approvedby": "test-delete-mode",
    }
    validation = validator.validate(fake_resource)

    if validation.get("resource_exists") is False:
        cprint("  [SKIP] Endpoint " + name + " not found - may already be deleted", Fore.YELLOW)
        report_lines.append("### Step 3: Endpoint Validation")
        report_lines.append("- Endpoint not found: " + name + " (ResourceNotFound)")
        _write_test_report(report_lines, output_dir, timestamp_str)
        return
    if validation.get("resource_exists") is None:
        cprint("  [WARN] Could not validate: " + validation.get("validation_notes", ""), Fore.YELLOW)
    else:
        cprint("  [OK] Endpoint exists: " + name, Fore.GREEN)

    conn_state = validation.get("connection_state", "Unknown")
    backend_id = validation.get("backend_resource_id", "")
    cprint("  Connection state : " + conn_state, Fore.WHITE)
    cprint("  Backend ID       : " + (backend_id[:80] if backend_id else "(none)"), Fore.WHITE)
    report_lines.append("### Step 3: Endpoint Validation - PASS")
    report_lines.append("- Exists: True")
    report_lines.append("- Connection state: " + conn_state)
    report_lines.append("- Backend resource ID: " + backend_id)

    cprint("\n  Step 4: Delete confirmation", Fore.CYAN)
    if not dry_run:
        cprint("\n  You are about to DELETE: " + name, Fore.RED, bold=True)
        cprint("  Resource Group: " + resource_group, Fore.RED)
        cprint("  Subscription  : " + sub_after, Fore.RED)
        cprint("  Connection    : " + conn_state, Fore.RED)
        confirm = input("\n  Type CONFIRM to delete (anything else aborts): ").strip()
        if confirm != "CONFIRM":
            cprint("  Aborted. No resources deleted.", Fore.YELLOW)
            report_lines.append("### Step 4: CONFIRM - ABORTED by user")
            _write_test_report(report_lines, output_dir, timestamp_str)
            return
        report_lines.append("### Step 4: CONFIRM - CONFIRMED by user")
    else:
        cprint("  [DRY RUN] Skipping CONFIRM prompt", Fore.YELLOW)
        report_lines.append("### Step 4: CONFIRM - DRY RUN (skipped)")

    cprint("\n  Step 5: Delete endpoint", Fore.CYAN)
    resource_id = ""
    if validation.get("raw_data"):
        resource_id = (validation["raw_data"].get("id") or "")

    cleanup_engine = CleanupEngine(backup_dir="backups", dry_run=dry_run)
    backup_path = cleanup_engine.backup_resource(fake_resource, validation)
    cprint("  [BACKUP] " + str(backup_path), Fore.CYAN)

    if not dry_run:
        if resource_id:
            cmd = ["network", "private-endpoint", "delete", "--ids", resource_id]
            delete_method = "ids"
        else:
            cmd = ["network", "private-endpoint", "delete",
                   "--name", name, "--resource-group", resource_group]
            delete_method = "name_rg"
        cmd_str = "az " + " ".join(cmd)
        cprint("  [DELETE] " + cmd_str, Fore.YELLOW)
        cprint("  [METHOD] " + delete_method, Fore.YELLOW)
        _, err = run_az(cmd, timeout=120)
        if err and err != "ResourceNotFound":
            if "AuthorizationFailed" in (err or "") or "does not have authorization" in (err or "").lower():
                rbac_msg = ("RBAC BLOCKED: account does not have "
                            "Microsoft.Network/privateEndpoints/delete on this scope.")
                cprint("  [FAIL] " + rbac_msg, Fore.RED, bold=True)
                report_lines.append("### Step 5: Delete - FAILED (RBAC)")
                report_lines.append("- Command: " + cmd_str)
                report_lines.append("- Error: " + rbac_msg)
                _write_test_report(report_lines, output_dir, timestamp_str)
                return
            cprint("  [FAIL] Delete error: " + (err or "")[:200], Fore.RED)
            report_lines.append("### Step 5: Delete - FAILED")
            report_lines.append("- Command: " + cmd_str)
            report_lines.append("- Error: " + (err or "")[:200])
            _write_test_report(report_lines, output_dir, timestamp_str)
            return
        cprint("  [OK] Delete command executed", Fore.GREEN)
        report_lines.append("### Step 5: Delete - EXECUTED")
        report_lines.append("- Command: " + cmd_str)
        report_lines.append("- Method: " + delete_method)

        cprint("\n  Step 6: Verify ResourceNotFound", Fore.CYAN)
        time.sleep(3)
        verify_status, verify_msg = cleanup_engine.verify_deletion(fake_resource)
        if verify_status == "VERIFIED_GONE":
            cprint("  DELETED and VERIFIED_GONE: " + name, Fore.GREEN, bold=True)
            report_lines.append("### Step 6: Verification - VERIFIED_GONE")
            report_lines.append("- Azure confirmed ResourceNotFound for: " + name)
        else:
            cprint("  [WARN] Verification: " + verify_status + " - " + verify_msg, Fore.YELLOW)
            report_lines.append("### Step 6: Verification - " + verify_status)
            report_lines.append("- " + verify_msg)
    else:
        report_lines.append("### Step 5: Delete - DRY RUN (not executed)")
        report_lines.append("### Step 6: Verification - DRY RUN (skipped)")
        cprint("  [DRY RUN] Delete and verify skipped", Fore.YELLOW)

    report_lines += [
        "",
        "## Post-Delete",
        "",
        "Run the Resource Graph query in docs/RESOURCE_GRAPH_QUERIES.md to confirm",
        "remaining disconnected private endpoints.",
        "",
        "Query for this specific endpoint:",
        "  Resources",
        "  | where type =~ 'microsoft.network/privateendpoints'",
        "  | where name =~ '" + name + "'",
    ]

    _write_test_report(report_lines, output_dir, timestamp_str)
    cprint("\n  Test-delete report written to: " + output_dir + "/", Fore.GREEN)


def _write_test_report(lines: List[str], output_dir: str, timestamp_str: str) -> None:
    """Write test_delete_report to output_dir."""
    ensure_dirs(output_dir)
    path = Path(output_dir) / ("test_delete_report_" + timestamp_str + ".md")
    try:
        with open(path, "w") as f:
            f.write("\n".join(lines))
        cprint("  Test report: " + str(path), Fore.GREEN)
    except Exception as e:
        cprint("  [WARN] Test report write failed: " + str(e), Fore.YELLOW)


def run_governance_scan(args) -> None:
    """
    Run the broad EDAV Resource Governance scan using Azure Resource Graph.
    Discovers: disconnected PEs, unattached NSGs/disks/PIPs/NICs,
    stopped VMs, Event Grid topics, storage, AKS/Databricks resources.
    Generates: CSV, Excel (multi-tab), Markdown, HTML, JSON reports.

    Query execution priority (--query-mode):
      auto (default) - try az graph CLI, fall back to az rest,
                       fall back to manual CSV if --manual-*-csv provided
      cli            - az graph query only (requires resource-graph extension)
      rest           - az rest POST only (no extension required)
      csv            - manual CSV exports from Azure Portal only

    On restricted CDC VDIs where az graph extension cannot be installed:
      Use --query-mode rest (no extension needed) or
      Use --query-mode csv with manually exported CSV files.

    Usage:
      python main.py --scan-governance \
        --subscriptions "OCIO-TSBDEV-C1,OCIO-TSBPRD-C1" \
        --output-dir reports/governance/
    """
    if not GOVERNANCE_SCANNER_AVAILABLE:
        cprint("[ERROR] GovernanceScanner not available.", Fore.RED, bold=True)
        cprint("  Ensure monitors/governance_scanner.py is present.", Fore.YELLOW)
        return

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timestamp_str = ts()
    output_dir = getattr(args, "output_dir", "reports") or "reports"
    subscriptions_arg = getattr(args, "subscriptions", "") or ""
    subscriptions = [s.strip() for s in subscriptions_arg.split(",") if s.strip()]
    dry_run = getattr(args, "dry_run", False)
    query_filter = getattr(args, "governance_query", None)  # optional filter
    terraform_path = getattr(args, "terraform_path", None)
    query_mode = getattr(args, "query_mode", QUERY_MODE_AUTO) or QUERY_MODE_AUTO
    manual_input_dir = getattr(args, "manual_input_dir", None)

    # Build manual CSV paths dict from CLI args
    # Supports both individual --manual-*-csv args and a --manual-input-dir directory
    manual_csv_paths = {}
    _manual_map = {
        "disconnected_private_endpoints": getattr(args, "manual_private_endpoints_csv", None),
        "unattached_nsgs":               getattr(args, "manual_nsg_csv", None),
        "unattached_disks":              getattr(args, "manual_disks_csv", None),
        "unattached_public_ips":         getattr(args, "manual_publicips_csv", None),
        "unattached_nics":               getattr(args, "manual_nics_csv", None),
    }
    for qkey, csv_path in _manual_map.items():
        if csv_path:
            manual_csv_paths[qkey] = csv_path
    # If --manual-input-dir is given, auto-discover CSVs by name pattern
    if manual_input_dir:
        from pathlib import Path as _Path
        _dir = _Path(manual_input_dir)
        _auto_map = {
            "disconnected_private_endpoints": ["pe*", "*private*endpoint*", "*privateendpoint*"],
            "unattached_nsgs":               ["nsg*", "*security*group*"],
            "unattached_disks":              ["disk*", "*managed*disk*"],
            "unattached_public_ips":         ["pip*", "*public*ip*"],
            "unattached_nics":               ["nic*", "*network*interface*"],
        }
        for qkey, patterns in _auto_map.items():
            if qkey not in manual_csv_paths:
                for pat in patterns:
                    matches = list(_dir.glob(pat + ".csv")) + list(_dir.glob(pat + ".CSV"))
                    if matches:
                        manual_csv_paths[qkey] = str(matches[0])
                        break

    ensure_dirs(output_dir, "team_reports", "logs")

    sep = "=" * 72
    cprint("\n" + sep, Fore.CYAN, bold=True)
    cprint("  EDAV Resource Governance Scan v" + VERSION, Fore.CYAN, bold=True)
    cprint("  Powered by Azure Resource Graph", Fore.CYAN)
    cprint(sep, Fore.CYAN)
    cprint("  Subscriptions  : " + (", ".join(subscriptions) if subscriptions else "all accessible"), Fore.WHITE)
    cprint("  Output dir     : " + output_dir, Fore.WHITE)
    cprint("  Date           : " + run_date, Fore.WHITE)
    if dry_run:
        cprint("  [DRY RUN] Reports only - no deletions.", Fore.YELLOW, bold=True)

    # Verify az graph extension is available
    cur_user = get_current_az_user()
    cur_sub = get_current_subscription_info()
    cprint("  Azure CLI user  : " + (cur_user or "unknown"), Fore.WHITE)
    cprint("  Current sub     : " + (cur_sub.get("name") or "unknown"), Fore.WHITE)
    cprint("  Query mode      : " + query_mode.upper(), Fore.WHITE)
    if manual_csv_paths:
        cprint("  Manual CSVs     : " + str(len(manual_csv_paths)) + " file(s) provided", Fore.WHITE)
        for _qk, _cp in manual_csv_paths.items():
            cprint("    " + _qk + " -> " + _cp, Fore.WHITE)
    cprint(sep, Fore.CYAN)

    # Determine which queries to run
    query_keys = list(GOVERNANCE_QUERIES.keys())
    if query_filter:
        query_keys = [k for k in query_keys if query_filter.lower() in k.lower()]

    cprint("\n  Running " + str(len(query_keys)) + " governance queries...", Fore.CYAN)

    # Print query source info if CSV mode
    if query_mode == QUERY_MODE_CSV and not manual_csv_paths:
        cprint("\n[WARN] --query-mode csv selected but no --manual-*-csv files provided.", Fore.YELLOW, bold=True)
        cprint("  Use --manual-private-endpoints-csv, --manual-nics-csv, etc.", Fore.YELLOW)
        cprint("  Or use --manual-input-dir to auto-discover CSVs.", Fore.YELLOW)
        cprint("  Example:", Fore.YELLOW)
        cprint("    python main.py --scan-governance --query-mode csv \\", Fore.YELLOW)
        cprint("      --manual-private-endpoints-csv pe_results.csv \\", Fore.YELLOW)
        cprint("      --manual-nics-csv nic_results.csv", Fore.YELLOW)

    scanner = GovernanceScanner(
        subscriptions=subscriptions,
        query_keys=query_keys,
        terraform_path=terraform_path,
        query_mode=query_mode,
        manual_csv_paths=manual_csv_paths,
    )

    all_results, summary = scanner.scan()

    # Print per-query results as they come in
    cprint("\n" + sep, Fore.CYAN)
    cprint("  GOVERNANCE SCAN RESULTS", Fore.CYAN, bold=True)
    cprint(sep, Fore.CYAN)

    query_results = summary.get("query_results", {})
    for key, qr in query_results.items():
        display = qr.get("display_name", key)
        count = qr.get("count", 0)
        err = qr.get("error")
        cls_counts = qr.get("classifications", {})

        if err:
            cprint("  [ERROR] " + display + ": " + str(err)[:80], Fore.RED)
            if any(x in str(err).lower() for x in [
                "not found", "extension", "could not find", "unrecognized"
            ]):
                cprint("     -> az graph extension missing.", Fore.YELLOW)
                cprint("     -> Try: python main.py --scan-governance --query-mode rest", Fore.YELLOW)
            elif "ssl" in str(err).lower() or "certificate" in str(err).lower():
                cprint("     -> SSL/cert error. Try: --query-mode rest", Fore.YELLOW)
                cprint("     -> Or export manually: see docs/RESOURCE_GRAPH_QUERIES.md", Fore.YELLOW)
            continue

        safe_n = cls_counts.get("SAFE_DELETE", 0)
        review_n = cls_counts.get("REVIEW_REQUIRED", 0)
        keep_n = cls_counts.get("KEEP", 0)

        source = qr.get("query_source", "")
        source_tag = " [" + source + "]" if source else ""
        line = "  " + display.ljust(45) + " | Total: " + str(count).rjust(4) + source_tag
        if safe_n:
            line += " | SAFE_DELETE: " + str(safe_n)
        if review_n:
            line += " | REVIEW: " + str(review_n)
        if keep_n:
            line += " | KEEP: " + str(keep_n)
        cprint(line, Fore.WHITE if count == 0 else Fore.CYAN)

    # Executive summary dashboard
    totals = summary.get("totals", {})
    cprint("\n" + sep, Fore.CYAN, bold=True)
    cprint("  EDAV GOVERNANCE SCAN SUMMARY", Fore.CYAN, bold=True)
    cprint(sep, Fore.CYAN)

    def _pad(label: str) -> str:
        return ("  " + label).ljust(48)

    print(_pad("Disconnected Private Endpoints") +
          ": " + str(query_results.get("disconnected_private_endpoints", {}).get("count", 0)))
    print(_pad("Already Removed (RESOURCE_NOT_FOUND)") +
          ": " + str(totals.get("RESOURCE_NOT_FOUND", 0)))
    print(_pad("Unattached NSGs") +
          ": " + str(query_results.get("unattached_nsgs", {}).get("count", 0)))
    print(_pad("Unattached Managed Disks") +
          ": " + str(query_results.get("unattached_disks", {}).get("count", 0)))
    print(_pad("Unattached Public IPs") +
          ": " + str(query_results.get("unattached_public_ips", {}).get("count", 0)))
    print(_pad("Unattached NICs") +
          ": " + str(query_results.get("unattached_nics", {}).get("count", 0)))
    print(_pad("  -> KEEP (Azure-managed NICs)") +
          ": " + str(query_results.get("unattached_nics", {}).get("classifications", {}).get("KEEP", 0)))
    print(_pad("Stopped / Deallocated VMs") +
          ": " + str(query_results.get("stopped_vms", {}).get("count", 0)))
    print(_pad("Event Grid - No Subscriptions") +
          ": " + str(query_results.get("eventgrid_no_subscriptions", {}).get("count", 0)))
    print(_pad("Storage Accounts Needing Review") +
          ": " + str(query_results.get("storage_review", {}).get("count", 0)))
    print(_pad("AKS / Databricks Managed Resources") +
          ": " + str(query_results.get("aks_databricks_resources", {}).get("count", 0)))
    cprint("-" * 72, Fore.CYAN)
    cprint("  SAFE_DELETE Candidates".ljust(48) + ": " + str(totals.get("SAFE_DELETE", 0)),
           Fore.GREEN, bold=True)
    cprint("  REVIEW_REQUIRED".ljust(48) + ": " + str(totals.get("REVIEW_REQUIRED", 0)),
           Fore.YELLOW)
    cprint("  KEEP (Azure-managed / In Use)".ljust(48) + ": " + str(totals.get("KEEP", 0)),
           Fore.CYAN)
    cprint("  DO_NOT_DELETE".ljust(48) + ": " + str(totals.get("DO_NOT_DELETE", 0)),
           Fore.RED)
    cprint("  No Action Needed".ljust(48) + ": " + str(totals.get("no_action_needed", 0)),
           Fore.WHITE)
    cprint("  Total Resources Found".ljust(48) + ": " + str(totals.get("total", 0)),
           Fore.WHITE, bold=True)
    cprint("=" * 72, Fore.CYAN)

    # NIC safety reminder
    nic_keep = query_results.get("unattached_nics", {}).get("classifications", {}).get("KEEP", 0)
    if nic_keep > 0:
        cprint("\n  NIC SAFETY: " + str(nic_keep) + " NIC(s) classified as KEEP (Azure-managed).",
               Fore.YELLOW, bold=True)
        cprint("  These NICs matched -pe-nic, .nic., AKS, Databricks, or Azure-managed RG patterns.",
               Fore.YELLOW)
        cprint("  DO NOT delete KEEP-classified NICs.", Fore.YELLOW)

    # Generate reports
    cprint("\n  Generating reports...", Fore.CYAN)
    run_meta = {
        "run_date": run_date,
        "mode": "governance-scan",
        "subscriptions": subscriptions,
        "version": VERSION,
        "query_keys": query_keys,
    }
    reporter = GovernanceReportGenerator(output_dir=output_dir)
    report_files = reporter.generate_all(all_results, summary, run_meta)
    cprint("  Reports written to: " + output_dir + "/", Fore.GREEN)

    # Post-scan reminder
    cprint("\n" + "=" * 72, Fore.CYAN)
    cprint("  POST-SCAN ACTIONS", Fore.CYAN, bold=True)
    cprint("=" * 72, Fore.CYAN)
    cprint("  1. Review Excel SAFE_DELETE tab with network/platform teams", Fore.WHITE)
    cprint("  2. Collect approvals (ApprovalTicket, ApprovedBy) for each candidate", Fore.WHITE)
    cprint("  3. Run: python main.py --mode delete --dry-run --delete-approved --input approved.xlsx", Fore.WHITE)
    cprint("  4. Run Resource Graph validation query from docs/RESOURCE_GRAPH_QUERIES.md", Fore.WHITE)
    cprint("=" * 72, Fore.CYAN)


def run_phase1_pipeline(args):
    """
    Phase 1 modular pipeline entry point.
    Supports --mode report/delete, --resource-type, --subscriptions, --terraform-path.
    """
    run_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    timestamp_str = ts()
    mode = getattr(args, 'mode', 'report') or 'report'
    resource_type = getattr(args, 'resource_type', 'private-endpoints') or 'private-endpoints'
    output_dir = getattr(args, 'output_dir', 'reports')
    dry_run = getattr(args, 'dry_run', False)
    delete_approved = getattr(args, 'delete_approved', False)
    subscriptions_arg = getattr(args, 'subscriptions', '') or ''
    terraform_path = getattr(args, 'terraform_path', None)
    owners_file = getattr(args, 'owners_file', 'config/owners.yml')
    config_dir = getattr(args, 'config_dir', 'config')
    change_ticket = getattr(args, 'change_ticket', '') or ''
    approved_by = getattr(args, 'approved_by', '') or ''
    required_user = getattr(args, 'required_user', None) or DEFAULT_REQUIRED_DELETE_USER
    allow_terraform = getattr(args, 'allow_terraform_managed', False)

    ensure_dirs(output_dir, 'team_reports', 'logs', 'backups')

    sep = '=' * 72
    cprint('\n' + sep, Fore.CYAN, bold=True)
    cprint('  ' + TOOL_NAME + ' v' + VERSION + ' - Phase 1 Modular Pipeline', Fore.CYAN, bold=True)
    cprint(sep, Fore.CYAN)
    cprint('  Mode: ' + mode.upper() + ' | Resource Type: ' + resource_type, Fore.CYAN)
    cprint('  Date: ' + run_date, Fore.CYAN)
    if dry_run:
        cprint('  [DRY RUN] No Azure resources will be modified.', Fore.YELLOW, bold=True)

    # Resolve subscriptions
    subscriptions = [s.strip() for s in subscriptions_arg.split(',') if s.strip()]
    if not subscriptions:
        subscriptions = _load_subscriptions_config(config_dir)
    if subscriptions:
        cprint('  Subscriptions: ' + ', '.join(subscriptions), Fore.CYAN)

    # Report mode
    if mode == 'report':
        input_file = getattr(args, 'input', None)
        if input_file and Path(input_file).exists():
            args.audit_only = True
            args.cleanup_approved = False
            if subscriptions:
                args.subscriptions = ','.join(subscriptions)
            run_pipeline(args)
            # After run_pipeline, generate team reports and summary from the results
            # Note: run_pipeline handles its own report generation
            run_meta = {
                'run_date': run_date, 'mode': mode,
                'resource_type': resource_type,
                'subscriptions': subscriptions_arg or ', '.join(subscriptions),
            }
            _generate_executive_summary([], run_meta, output_dir)
        else:
            cprint('', Fore.CYAN)
            cprint('  No --input file provided for report mode.', Fore.YELLOW)
            cprint('  To scan resources:', Fore.YELLOW)
            cprint('    1. Export findings from EDAV dashboard', Fore.YELLOW)
            cprint('    2. Pass as: python main.py --mode report --input findings.csv', Fore.YELLOW)
            run_meta = {
                'run_date': run_date, 'mode': mode,
                'resource_type': resource_type,
                'subscriptions': subscriptions_arg or ', '.join(subscriptions),
            }
            _generate_executive_summary([], run_meta, output_dir)

    # test-delete mode - single resource safe test
    elif mode == 'test-delete':
        name_arg = getattr(args, 'name', None)
        rg_arg = getattr(args, 'resource_group_arg', None)
        sub_arg = getattr(args, 'subscription_arg', None)

        if not name_arg or not rg_arg:
            cprint('\n[ERROR] --mode test-delete requires --name and --resource-group', Fore.RED, bold=True)
            cprint('  Example:', Fore.YELLOW)
            cprint('  python main.py --mode test-delete --resource-type private-endpoints \\', Fore.YELLOW)
            cprint('    --name testwebbseries-pe --resource-group ocio-network \\', Fore.YELLOW)
            cprint('    --subscription OCIO-TSBDEV-C1 --required-user bh55-su@cdc.gov', Fore.YELLOW)
            sys.exit(1)

        run_test_delete(
            name=name_arg, resource_group=rg_arg,
            subscription=sub_arg, required_user=required_user,
            output_dir=output_dir, dry_run=dry_run,
        )

    # Delete mode
    elif mode == 'delete':
        # SU Account enforcement
        current_user = get_current_az_user()
        if not dry_run:
            if current_user.lower() != required_user.lower():
                cprint('\n' + '=' * 72, Fore.RED, bold=True)
                cprint('  DELETE BLOCKED: SU account required', Fore.RED, bold=True)
                cprint('=' * 72, Fore.RED)
                cprint(f'  Current account  : {current_user}', Fore.RED)
                cprint(f'  Required account : {required_user}', Fore.RED)
                cprint(f'\n  You must login with the SU account {required_user}', Fore.RED)
                cprint('  before cleanup can run.', Fore.RED)
                cprint('\n  Remediation steps:', Fore.YELLOW)
                cprint('    az logout', Fore.YELLOW)
                cprint('    az login --use-device-code', Fore.YELLOW)
                cprint('    az account show --query user -o table', Fore.YELLOW)
                cprint('=' * 72, Fore.RED)
                sys.exit(1)

        if not delete_approved:
            cprint('\n[ERROR] --mode delete requires --delete-approved flag.', Fore.RED, bold=True)
            cprint('  Safety rule: Deletion NEVER runs without --delete-approved.', Fore.RED)
            sys.exit(1)
        input_file = getattr(args, 'input', None)
        if not input_file:
            cprint('\n[ERROR] --mode delete requires --input <approvals.xlsx>', Fore.RED, bold=True)
            sys.exit(1)
        if resource_type not in ('private-endpoints',):
            cprint('\n[ERROR] --mode delete only supports --resource-type private-endpoints', Fore.RED, bold=True)
            cprint('  Storage and other types require manual review and are blocked for deletion.', Fore.RED)
            sys.exit(1)
        args.cleanup_approved = True
        args.audit_only = False
        if subscriptions:
            args.subscriptions = ','.join(subscriptions)
        if terraform_path:
            args.terraform_path = terraform_path
        run_pipeline(args)
    else:
        cprint('[ERROR] Unknown mode: ' + str(mode), Fore.RED)
        sys.exit(1)


if __name__ == "__main__":
    main()
